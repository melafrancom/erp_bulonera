"""
Exportadores de Estado de Cuenta (Mayor) a Excel y PDF para la app Customers.
"""
import io
import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors

from common.company import get_company_info
from common.utils import format_currency, format_cuit


def export_account_statement_excel(statement_data: dict) -> io.BytesIO:
    """
    Genera un archivo Excel (.xlsx) con el estado de cuenta (mayor) del cliente.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Cuenta"

    # Mostrar líneas de cuadrícula
    ws.views.sheetView[0].showGridLines = True

    customer = statement_data.get('customer')
    company = get_company_info()

    # Estilos
    TITLE_FONT = Font(name='Segoe UI', size=16, bold=True, color='1F4E78')
    SUBTITLE_FONT = Font(name='Segoe UI', size=11, bold=True, color='595959')
    LABEL_FONT = Font(name='Segoe UI', size=10, bold=True)
    VALUE_FONT = Font(name='Segoe UI', size=10)

    HEADER_FILL = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    HEADER_FONT = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')

    TOTAL_FILL = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    TOTAL_FONT = Font(name='Segoe UI', size=10, bold=True)

    BORDER_THIN = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Cabecera institucional
    ws['A1'] = company['name']
    ws['A1'].font = TITLE_FONT

    ws['A2'] = f"CUIT: {company['cuit']} - ESTADO DE CUENTA DE CLIENTE (MAYOR)"
    ws['A2'].font = SUBTITLE_FONT

    # 2. Información del cliente
    ws['A4'] = "Cliente:"
    ws['A4'].font = LABEL_FONT
    ws['B4'] = customer.business_name if customer else "N/A"
    ws['B4'].font = VALUE_FONT

    ws['D4'] = "CUIT / CUIL:"
    ws['D4'].font = LABEL_FONT
    ws['E4'] = format_cuit(customer.cuit_cuil) if customer and customer.cuit_cuil else "S/D"
    ws['E4'].font = VALUE_FONT

    ws['A5'] = "Condición IVA:"
    ws['A5'].font = LABEL_FONT
    ws['B5'] = customer.get_tax_condition_display() if customer else "N/A"
    ws['B5'].font = VALUE_FONT

    ws['D5'] = "Modalidad CC:"
    ws['D5'].font = LABEL_FONT
    ws['E5'] = customer.get_account_modality_display() if customer else "N/A"
    ws['E5'].font = VALUE_FONT

    # Resumen de saldos
    ws['A7'] = "Límite Crédito:"
    ws['A7'].font = LABEL_FONT
    ws['B7'] = float(statement_data.get('credit_limit', 0))
    ws['B7'].number_format = '$#,##0.00'

    ws['C7'] = "Deuda Total:"
    ws['C7'].font = LABEL_FONT
    ws['D7'] = float(statement_data.get('deuda_total', 0))
    ws['D7'].number_format = '$#,##0.00'

    ws['E7'] = "Disponible:"
    ws['E7'].font = LABEL_FONT
    ws['F7'] = float(statement_data.get('credito_disponible', 0))
    ws['F7'].number_format = '$#,##0.00'

    # Rango de fechas
    df_str = statement_data.get('date_from').strftime('%d/%m/%Y') if statement_data.get('date_from') else "Inicio"
    dt_str = statement_data.get('date_to').strftime('%d/%m/%Y') if statement_data.get('date_to') else "Hoy"
    ws['A8'] = f"Período: {df_str} a {dt_str}"
    ws['A8'].font = Font(name='Segoe UI', size=9, italic=True, color='7F7F7F')

    # 3. Tabla de movimientos
    headers = ['Fecha', 'Tipo', 'Comprobante', 'Debe ($)', 'Haber ($)', 'Saldo ($)']
    start_row = 10

    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=h_text)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    current_row = start_row + 1

    # Fila de Saldo Inicial si aplica
    initial_bal = statement_data.get('initial_balance', Decimal('0.00'))
    if initial_bal != Decimal('0.00') or statement_data.get('date_from'):
        ws.cell(row=current_row, column=1, value="-").alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=2, value="Saldo Anterior")
        ws.cell(row=current_row, column=3, value="Arrastre de Período")
        c_debe = ws.cell(row=current_row, column=4, value=float(initial_bal) if initial_bal > 0 else 0.0)
        c_debe.number_format = '$#,##0.00'
        c_haber = ws.cell(row=current_row, column=5, value=float(-initial_bal) if initial_bal < 0 else 0.0)
        c_haber.number_format = '$#,##0.00'
        c_saldo = ws.cell(row=current_row, column=6, value=float(initial_bal))
        c_saldo.number_format = '$#,##0.00'
        c_saldo.font = Font(bold=True)

        for col_num in range(1, 7):
            ws.cell(row=current_row, column=col_num).border = BORDER_THIN

        current_row += 1

    # Movimientos
    movements = statement_data.get('movements', [])
    for m in movements:
        m_date = m['date'].strftime('%d/%m/%Y') if hasattr(m['date'], 'strftime') else str(m['date'])

        ws.cell(row=current_row, column=1, value=m_date).alignment = Alignment(horizontal='center')
        ws.cell(row=current_row, column=2, value=m.get('type_display', ''))
        ws.cell(row=current_row, column=3, value=m.get('comprobante', ''))

        c_debe = ws.cell(row=current_row, column=4, value=float(m.get('debe', 0)))
        c_debe.number_format = '$#,##0.00'
        c_debe.alignment = Alignment(horizontal='right')

        c_haber = ws.cell(row=current_row, column=5, value=float(m.get('haber', 0)))
        c_haber.number_format = '$#,##0.00'
        c_haber.alignment = Alignment(horizontal='right')

        c_saldo = ws.cell(row=current_row, column=6, value=float(m.get('saldo', 0)))
        c_saldo.number_format = '$#,##0.00'
        c_saldo.alignment = Alignment(horizontal='right')

        for col_num in range(1, 7):
            ws.cell(row=current_row, column=col_num).border = BORDER_THIN

        current_row += 1

    # Fila de Totales
    ws.cell(row=current_row, column=1, value="TOTALES").alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=1).font = TOTAL_FONT
    ws.cell(row=current_row, column=1).fill = TOTAL_FILL

    for c_idx in [2, 3]:
        cell = ws.cell(row=current_row, column=c_idx, value="")
        cell.fill = TOTAL_FILL

    tot_debe = ws.cell(row=current_row, column=4, value=float(statement_data.get('total_debe', 0)))
    tot_debe.number_format = '$#,##0.00'
    tot_debe.font = TOTAL_FONT
    tot_debe.fill = TOTAL_FILL

    tot_haber = ws.cell(row=current_row, column=5, value=float(statement_data.get('total_haber', 0)))
    tot_haber.number_format = '$#,##0.00'
    tot_haber.font = TOTAL_FONT
    tot_haber.fill = TOTAL_FILL

    tot_saldo = ws.cell(row=current_row, column=6, value=float(statement_data.get('saldo_final', 0)))
    tot_saldo.number_format = '$#,##0.00'
    tot_saldo.font = TOTAL_FONT
    tot_saldo.fill = TOTAL_FILL

    for col_num in range(1, 7):
        ws.cell(row=current_row, column=col_num).border = BORDER_THIN

    # Auto-ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_account_statement_pdf(statement_data: dict) -> io.BytesIO:
    """
    Genera un documento PDF (.pdf) con el estado de cuenta (mayor) del cliente usando ReportLab.
    """
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    PAGE_W, PAGE_H = A4
    MARGIN = 15 * mm

    customer = statement_data.get('customer')
    company = get_company_info()

    # 1. Cabecera institucional
    c.setFont("Helvetica-Bold", 14)
    c.drawString(MARGIN, PAGE_H - MARGIN - 10, company['name'])

    c.setFont("Helvetica", 9)
    c.drawString(MARGIN, PAGE_H - MARGIN - 23, f"CUIT: {company['cuit']}")

    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(PAGE_W - MARGIN, PAGE_H - MARGIN - 10, "ESTADO DE CUENTA CORRIENTE")

    c.setFont("Helvetica", 9)
    fecha_emision = datetime.date.today().strftime('%d/%m/%Y')
    c.drawRightString(PAGE_W - MARGIN, PAGE_H - MARGIN - 23, f"Fecha Emisión: {fecha_emision}")

    c.setLineWidth(0.5)
    c.line(MARGIN, PAGE_H - MARGIN - 30, PAGE_W - MARGIN, PAGE_H - MARGIN - 30)

    # 2. Caja de información del cliente
    box_y = PAGE_H - MARGIN - 38
    c.setFont("Helvetica-Bold", 10)
    c.drawString(MARGIN, box_y, f"Cliente: {customer.business_name if customer else 'N/A'}")

    c.setFont("Helvetica", 9)
    cuit_str = format_cuit(customer.cuit_cuil) if customer and customer.cuit_cuil else "S/D"
    c.drawString(MARGIN, box_y - 14, f"CUIT: {cuit_str}  |  Cond. IVA: {customer.get_tax_condition_display() if customer else 'N/A'}")

    df_str = statement_data.get('date_from').strftime('%d/%m/%Y') if statement_data.get('date_from') else "Inicio"
    dt_str = statement_data.get('date_to').strftime('%d/%m/%Y') if statement_data.get('date_to') else "Hoy"
    c.drawString(MARGIN, box_y - 26, f"Período consultado: {df_str} al {dt_str}")

    # Resumen financiero a la derecha
    c.setFont("Helvetica-Bold", 9)
    deuda_str = format_currency(statement_data.get('deuda_total', 0))
    disp_str = format_currency(statement_data.get('credito_disponible', 0))
    limite_str = format_currency(statement_data.get('credit_limit', 0))

    c.drawRightString(PAGE_W - MARGIN, box_y, f"Deuda Actual: {deuda_str}")
    c.setFont("Helvetica", 9)
    c.drawRightString(PAGE_W - MARGIN, box_y - 14, f"Disponible: {disp_str}")
    c.drawRightString(PAGE_W - MARGIN, box_y - 26, f"Límite Crédito: {limite_str}")

    c.line(MARGIN, box_y - 33, PAGE_W - MARGIN, box_y - 33)

    # 3. Tabla de movimientos
    table_top = box_y - 45
    headers = ['Fecha', 'Concepto / Tipo', 'Comprobante', 'Debe ($)', 'Haber ($)', 'Saldo ($)']

    table_data = [headers]

    initial_bal = statement_data.get('initial_balance', Decimal('0.00'))
    if initial_bal != Decimal('0.00') or statement_data.get('date_from'):
        table_data.append([
            '-',
            'Saldo Anterior',
            'Arrastrado',
            format_currency(initial_bal) if initial_bal > 0 else "$ 0,00",
            format_currency(-initial_bal) if initial_bal < 0 else "$ 0,00",
            format_currency(initial_bal)
        ])

    movements = statement_data.get('movements', [])
    for m in movements:
        m_date = m['date'].strftime('%d/%m/%Y') if hasattr(m['date'], 'strftime') else str(m['date'])
        table_data.append([
            m_date,
            m.get('type_display', ''),
            m.get('comprobante', ''),
            format_currency(m.get('debe', 0)),
            format_currency(m.get('haber', 0)),
            format_currency(m.get('saldo', 0))
        ])

    # Totales
    table_data.append([
        'TOTALES',
        '',
        '',
        format_currency(statement_data.get('total_debe', 0)),
        format_currency(statement_data.get('total_haber', 0)),
        format_currency(statement_data.get('saldo_final', 0))
    ])

    content_w = PAGE_W - (2 * MARGIN)
    col_widths = [22 * mm, 40 * mm, 50 * mm, 23 * mm, 23 * mm, 23 * mm]

    t = Table(table_data, colWidths=col_widths)
    t_style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E78')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D9D9D9')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        # Totales row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E7E6E6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]

    t.setStyle(TableStyle(t_style))
    t.wrapOn(c, content_w, PAGE_H)
    t_h = t._height

    # Renderizar la tabla (si sobrepasa página simple, dibujamos en página actual)
    t.drawOn(c, MARGIN, table_top - t_h)

    # Pie de página
    c.setFont("Helvetica-Oblique", 8)
    c.drawString(MARGIN, 10 * mm, f"Documento generado automáticamente por BULONERA ERP — {company['name']}")

    c.showPage()
    c.save()

    buf.seek(0)
    return buf
