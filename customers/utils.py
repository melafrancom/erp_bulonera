from django.contrib.auth import get_user_model
import pandas as pd
from django.db import transaction
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from decimal import Decimal
import re

# Local apps
from .models import Customer, CustomerSegment

User = get_user_model()


class CustomerExcelManager:
    """
    Manager for handling customer data import/export from Excel files.
    """
    
    def __init__(self, user=None, db_alias='default'):
        self.user = user
        self.db_alias = db_alias
        self.validation_errors = []
    
    def import_customer_data(self, file_path_or_buffer, update_existing_customers=True):
        """
        Import customers from an Excel file with validation.
        """
        results = {
            'created_customers': 0,
            'updated_customers': 0,
            'validation_errors': [],
            'skipped_rows': 0
        }
        
        try:
            # Parse Excel file
            customers_data = self.parse_excel(file_path_or_buffer)
            if not customers_data:
                results['validation_errors'].extend(self.validation_errors)
                return results
            
            # Process data atomically
            with transaction.atomic(using=self.db_alias):
                for i, item in enumerate(customers_data):
                    try:
                        row_number = i + 2  # Considering Excel header
                        
                        # Validate required fields
                        if not self._validate_required_fields(item, row_number):
                            results['skipped_rows'] += 1
                            continue
                        
                        # Process customer
                        customer, created = self._process_customer(item, update_existing_customers)
                        if not customer:
                            results['skipped_rows'] += 1
                            continue
                        
                        if created:
                            results['created_customers'] += 1
                        else:
                            results['updated_customers'] += 1
                        
                        # Process segment relationship
                        self._process_segment(customer, item)
                    
                    except Exception as e:
                        self.validation_errors.append({
                            'row': row_number,
                            'field': 'general',
                            'error': f"Error procesando cliente CUIT {item.get('cuit_cuil', 'desconocido')}: {str(e)}"
                        })
                        results['skipped_rows'] += 1
        
        except Exception as e:
            self.validation_errors.append({
                'row': 'N/A',
                'field': 'file',
                'error': f"Error general en la importación: {str(e)}"
            })
        
        results['validation_errors'].extend(self.validation_errors)
        return results
    
    def _validate_required_fields(self, item, row_number):
        """
        Validate required fields.
        """
        required_fields = ['cuit_cuil', 'business_name']
        
        for field in required_fields:
            value = item.get(field)
            if not value or (isinstance(value, str) and not value.strip()):
                self.validation_errors.append({
                    'row': row_number,
                    'field': field,
                    'error': f"El campo '{field}' es obligatorio y no puede estar vacío."
                })
                return False
        
        # Validate CUIT format
        cuit = str(item.get('cuit_cuil')).strip()
        if not self._validate_cuit_format(cuit):
            self.validation_errors.append({
                'row': row_number,
                'field': 'cuit_cuil',
                'error': f"El CUIT '{cuit}' no tiene un formato válido (debe ser XX-XXXXXXXX-X)."
            })
            return False
        
        return True
    
    def _validate_cuit_format(self, cuit):
        """
        Validate CUIT/CUIL format: XX-XXXXXXXX-X
        """
        # Remove any spaces
        cuit = cuit.replace(' ', '')
        
        # If already has dashes, validate format
        if '-' in cuit:
            pattern = r'^\d{2}-\d{8}-\d{1}$'
            return bool(re.match(pattern, cuit))
        
        # If no dashes, check if it's 11 digits
        if len(cuit) == 11 and cuit.isdigit():
            return True
        
        return False
    
    def _format_cuit(self, cuit):
        """
        Format CUIT to XX-XXXXXXXX-X if needed.
        """
        cuit = str(cuit).replace(' ', '').replace('-', '')
        
        if len(cuit) == 11 and cuit.isdigit():
            return f"{cuit[0:2]}-{cuit[2:10]}-{cuit[10]}"
        
        return cuit
    
    def _process_customer(self, item, update_existing=True):
        """
        Process and save/update a customer.
        """
        cuit = self._format_cuit(item.get('cuit_cuil'))
        
        # Try to get existing customer by CUIT
        customer = Customer.objects.using(self.db_alias).filter(cuit_cuil=cuit).first()
        created = False
        
        if customer:
            if not update_existing:
                return customer, False
        else:
            customer = Customer()
            created = True
        
        # Set basic fields
        customer.cuit_cuil = cuit
        customer.business_name = item.get('business_name', '').strip()
        customer.trade_name = item.get('trade_name', '').strip() if item.get('trade_name') else ''
        
        # Customer type
        customer_type = str(item.get('customer_type', 'PERSON')).upper()
        if customer_type in ['PERSONA', 'PERSON', 'FISICA']:
            customer.customer_type = 'PERSON'
        elif customer_type in ['EMPRESA', 'COMPANY']:
            customer.customer_type = 'COMPANY'
        
        # Tax condition mapping
        tax_condition = str(item.get('tax_condition', 'CF')).upper()
        tax_mapping = {
            'RI': 'RI', 'RESPONSABLE': 'RI', 'INSCRIPTO': 'RI',
            'MONO': 'MONO', 'MONOTRIBUTO': 'MONO', 'MONOTRIBUTISTA': 'MONO',
            'EX': 'EX', 'EXENTO': 'EX',
            'CF': 'CF', 'CONSUMIDOR': 'CF', 'FINAL': 'CF',
            'NR': 'NR',
        }
        customer.tax_condition = tax_mapping.get(tax_condition, 'CF')
        
        # Contact information
        customer.email = item.get('email', '').strip() if item.get('email') else ''
        customer.phone = item.get('phone', '').strip() if item.get('phone') else ''
        customer.mobile = item.get('mobile', '').strip() if item.get('mobile') else ''
        customer.website = item.get('website', '').strip() if item.get('website') else ''
        customer.contact_person = item.get('contact_person', '').strip() if item.get('contact_person') else ''
        
        # Billing address
        customer.billing_address = item.get('billing_address', '').strip() if item.get('billing_address') else ''
        customer.billing_city = item.get('billing_city', '').strip() if item.get('billing_city') else ''
        customer.billing_state = item.get('billing_state', '').strip() if item.get('billing_state') else ''
        customer.billing_zip_code = item.get('billing_zip_code', '').strip() if item.get('billing_zip_code') else ''
        customer.billing_country = item.get('billing_country', 'Argentina').strip()
        
        # Commercial terms
        customer.payment_term = int(item.get('payment_term', 0)) if item.get('payment_term') else 0
        customer.credit_limit = Decimal(str(item.get('credit_limit', 0))) if item.get('credit_limit') else Decimal('0.00')
        customer.discount_percentage = Decimal(str(item.get('discount_percentage', 0))) if item.get('discount_percentage') else Decimal('0.00')
        
        # Flags
        allow_credit = str(item.get('allow_credit', '')).strip().upper() if item.get('allow_credit') else ''
        customer.allow_credit = allow_credit in ['SI', 'SÍ', 'YES', 'TRUE', '1', 'VERDADERO']
        
        # Notes
        customer.notes = item.get('notes', '').strip() if item.get('notes') else ''
        
        # Set audit fields
        if created and self.user:
            customer.created_by = self.user
        if self.user:
            customer.updated_by = self.user
        
        customer.save(using=self.db_alias)
        return customer, created
    
    def _process_segment(self, customer, item):
        """
        Process and assign customer segment.
        """
        segment_name = item.get('customer_segment', '').strip() if item.get('customer_segment') else ''
        
        if segment_name:
            # Get or create segment
            segment, _ = CustomerSegment.objects.using(self.db_alias).get_or_create(
                name=segment_name,
                defaults={
                    'description': f'Segmento creado automáticamente: {segment_name}',
                    'created_by': self.user if self.user else None,
                }
            )
            customer.customer_segment = segment
            customer.save(using=self.db_alias, update_fields=['customer_segment'])
    
    def parse_excel(self, file):
        """
        Parse Excel file and return list of dictionaries.
        """
        try:
            # Read Excel with pandas
            df = pd.read_excel(file, engine='openpyxl')
            
            # Replace NaN with None
            df = df.where(pd.notnull(df), None)
            
            # Clean column names
            # Expected format: lower case, spaces to underscores
            # Handling: "CUIT/CUIL *" -> "cuit_cuil"
            
            def clean_header(header):
                if not header:
                    return ''
                h = str(header).lower().strip()
                # Remove common symbols in templates
                h = h.replace('*', '').strip()
                # Handle specific mappings
                if 'cuit' in h:
                    return 'cuit_cuil'
                if 'razón social' in h or 'razon social' in h:
                    return 'business_name'
                # Standard cleanup
                h = h.replace(' ', '_').replace('/', '_')
                return h

            df.columns = [clean_header(c) for c in df.columns]
            
            # Remove empty rows (all NaNs)
            df.dropna(how='all', inplace=True)
            
            return df.to_dict('records')
        
        except Exception as e:
            self.validation_errors.append({
                'row': 'N/A',
                'field': 'file',
                'error': f"Error al leer el archivo Excel: {str(e)}"
            })
            return []
    
    def export_customers_to_excel(self):
        """
        Export customers to an Excel file.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Define headers
        headers = [
            'CUIT/CUIL', 'Razón Social', 'Nombre Comercial', 'Tipo Cliente',
            'Condición IVA', 'Email', 'Teléfono', 'Celular',
            'Dirección', 'Ciudad', 'Provincia', 'CP',
            'Segmento', 'Plazo Pago (días)', 'Límite Crédito',
            'Descuento %', 'Permite Crédito', 'Activo'
        ]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Get data
        customers = Customer.objects.using(self.db_alias).filter(is_active=True).select_related('customer_segment')
        
        # Write rows
        for row_num, customer in enumerate(customers, 2):
            ws.cell(row=row_num, column=1).value = customer.cuit_cuil
            ws.cell(row=row_num, column=2).value = customer.business_name
            ws.cell(row=row_num, column=3).value = customer.trade_name
            ws.cell(row=row_num, column=4).value = customer.get_customer_type_display()
            ws.cell(row=row_num, column=5).value = customer.get_tax_condition_display()
            ws.cell(row=row_num, column=6).value = customer.email
            ws.cell(row=row_num, column=7).value = customer.phone
            ws.cell(row=row_num, column=8).value = customer.mobile
            ws.cell(row=row_num, column=9).value = customer.billing_address
            ws.cell(row=row_num, column=10).value = customer.billing_city
            ws.cell(row=row_num, column=11).value = customer.billing_state
            ws.cell(row=row_num, column=12).value = customer.billing_zip_code
            ws.cell(row=row_num, column=13).value = customer.customer_segment.name if customer.customer_segment else ''
            ws.cell(row=row_num, column=14).value = customer.payment_term
            ws.cell(row=row_num, column=15).value = float(customer.credit_limit)
            ws.cell(row=row_num, column=16).value = float(customer.discount_percentage)
            ws.cell(row=row_num, column=17).value = 'Sí' if customer.allow_credit else 'No'
            ws.cell(row=row_num, column=18).value = 'Sí' if customer.is_active else 'No'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def generate_import_template(self):
        """
        Generate an Excel template for customer import.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla Clientes"
        
        headers = [
            ('cuit_cuil', 'CUIT/CUIL *', 'XX-XXXXXXXX-X'),
            ('business_name', 'Razón Social *', 'Nombre legal o completo'),
            ('trade_name', 'Nombre Comercial', 'Nombre fantasía (opcional)'),
            ('customer_type', 'Tipo Cliente', 'PERSON o COMPANY'),
            ('tax_condition', 'Condición IVA', 'RI, MONO, CF, EX, NR'),
            ('email', 'Email', 'email@ejemplo.com'),
            ('phone', 'Teléfono', '0291-4XXXXXX'),
            ('mobile', 'Celular', '291-15XXXXXX'),
            ('contact_person', 'Persona Contacto', 'Nombre del contacto'),
            ('billing_address', 'Dirección', 'Calle y número'),
            ('billing_city', 'Ciudad', 'Ciudad'),
            ('billing_state', 'Provincia', 'Provincia/Estado'),
            ('billing_zip_code', 'Código Postal', 'XXXX'),
            ('customer_segment', 'Segmento', 'Mayorista, Minorista, etc'),
            ('payment_term', 'Plazo Pago (días)', '0 = contado'),
            ('credit_limit', 'Límite Crédito', '0.00'),
            ('discount_percentage', 'Descuento %', '0.00'),
            ('allow_credit', 'Permite Crédito', 'SI o NO'),
            ('notes', 'Observaciones', 'Notas adicionales')
        ]
        
        for col_num, (field, display, help_text) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = display
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            help_cell = ws.cell(row=2, column=col_num)
            help_cell.value = help_text
            help_cell.font = Font(italic=True, size=9, color="666666")
            
        # Example row
        example_data = [
            '20-12345678-9', 'Juan Pérez', 'Ferretería El Tornillo', 'PERSON', 'MONO',
            'juan@email.com', '0291-4567890', '291-151234567', 'Juan Pérez',
            'Av. Colón 123', 'Bahía Blanca', 'Buenos Aires', '8000',
            'Minorista', '0', '0.00', '0.00', 'NO', 'Cliente de ejemplo'
        ]
        
        for col_num, value in enumerate(example_data, 1):
            ws.cell(row=3, column=col_num).value = value
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 18
            
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
