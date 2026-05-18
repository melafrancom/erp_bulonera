"""
ExportService: Exportación de reportes financieros a Excel y PDF.

Maneja la generación de archivos exportables a partir de los datos
calculados por ProfitAndLossService y CashFlowService.
"""
from io import BytesIO
from datetime import date
from decimal import Decimal
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .pnl_service import ProfitAndLossService
from .cashflow_service import CashFlowService

logger = logging.getLogger('reports')


class ExportService:
    """Genera archivos exportables (Excel, PDF) a partir de datos financieros."""

    # Estilos para Excel
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SECTION_FONT = Font(bold=True, size=10)
    
    TOTAL_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=10)
    
    SUBTOTAL_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    @staticmethod
    def export_pnl_to_xlsx(date_from: date, date_to: date) -> bytes:
        """
        Genera Excel con Estado de Resultados formateado.

        Estructura:
          Hoja 1: "Estado de Resultados"
            - Encabezado: Bulonera Alvear | Período
            - Tabla: Revenue, COGS, Gross Profit, OPEX (desglosado), EBITDA
            - Márgenes en %

          Hoja 2: "Detalle OPEX"
            - Desglose completo por categoría de gasto

        Args:
            date_from: Fecha inicial
            date_to: Fecha final

        Returns:
            bytes del archivo .xlsx
        """
        wb = Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # Calcular datos P&L
        pnl_service = ProfitAndLossService()
        pnl_data = pnl_service.get_pnl(date_from, date_to)

        # Hoja 1: Estado de Resultados
        ws1 = wb.create_sheet("Estado de Resultados", 0)
        ExportService._write_pnl_sheet(ws1, pnl_data, date_from, date_to)

        # Hoja 2: Detalle OPEX
        ws2 = wb.create_sheet("Detalle OPEX", 1)
        ExportService._write_opex_detail_sheet(ws2, pnl_data)

        # Retornar como bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_cashflow_to_xlsx(date_from: date, date_to: date) -> bytes:
        """
        Genera Excel del Flujo de Caja.

        Estructura:
          Hoja 1: "Flujo de Caja"
            - Inflows por método de pago
            - Outflows (gastos pagados)
            - Flujo neto

        Args:
            date_from: Fecha inicial
            date_to: Fecha final

        Returns:
            bytes del archivo .xlsx
        """
        wb = Workbook()
        wb.remove(wb.active)

        # Calcular datos CashFlow
        cf_service = CashFlowService()
        cf_data = cf_service.get_cashflow(date_from, date_to)

        # Hoja 1: Flujo de Caja
        ws = wb.create_sheet("Flujo de Caja", 0)
        ExportService._write_cashflow_sheet(ws, cf_data, date_from, date_to)

        # Retornar como bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def _write_pnl_sheet(ws, pnl_data: dict, date_from: date, date_to: date) -> None:
        """Escribe los datos del P&L en la hoja de trabajo."""
        
        # Configurar ancho de columnas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18

        row = 1

        # Encabezado
        ws[f'A{row}'] = "BULONERA ALVEAR"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        ws[f'A{row}'] = f"Estado de Resultados"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = f"Período: {date_from} a {date_to}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2

        # Encabezado de tabla
        ws[f'A{row}'] = "CONCEPTO"
        ws[f'B{row}'] = "MONTO ($)"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.fill = ExportService.HEADER_FILL
            cell.font = ExportService.HEADER_FONT
            cell.border = ExportService.BORDER
            cell.alignment = Alignment(horizontal='center')
        row += 1

        # INGRESOS
        ws[f'A{row}'] = "Ventas Brutas"
        ws[f'B{row}'] = pnl_data['revenue']['gross_sales']
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

        ws[f'A{row}'] = "Notas de Crédito"
        ws[f'B{row}'] = -pnl_data['revenue']['credit_notes']
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

        # INGRESOS NETOS
        ws[f'A{row}'] = "INGRESOS NETOS"
        ws[f'B{row}'] = pnl_data['revenue']['net_revenue']
        ws[f'A{row}'].fill = ExportService.SECTION_FILL
        ws[f'B{row}'].fill = ExportService.SECTION_FILL
        ws[f'A{row}'].font = ExportService.SECTION_FONT
        ws[f'B{row}'].font = ExportService.SECTION_FONT
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2

        # COGS
        ws[f'A{row}'] = "(-) COGS (Costo de Bienes Vendidos)"
        ws[f'B{row}'] = -pnl_data['cogs']
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

        # MARGEN BRUTO
        ws[f'A{row}'] = "MARGEN BRUTO"
        ws[f'B{row}'] = pnl_data['gross_profit']
        ws[f'A{row}'].fill = ExportService.SECTION_FILL
        ws[f'B{row}'].fill = ExportService.SECTION_FILL
        ws[f'A{row}'].font = ExportService.SECTION_FONT
        ws[f'B{row}'].font = ExportService.SECTION_FONT
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

        ws[f'A{row}'] = "Margen Bruto %"
        ws[f'B{row}'] = pnl_data['gross_margin_pct'] / 100
        ws[f'B{row}'].number_format = '0.00%'
        row += 2

        # OPEX
        ws[f'A{row}'] = "GASTOS OPERATIVOS"
        ws[f'A{row}'].fill = ExportService.SUBTOTAL_FILL
        ws[f'A{row}'].font = Font(bold=True, size=10)
        row += 1

        for category, amount in pnl_data['opex']['by_category'].items():
            ws[f'A{row}'] = f"  {category}"
            ws[f'B{row}'] = -amount
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1

        ws[f'A{row}'] = "TOTAL OPEX"
        ws[f'B{row}'] = -pnl_data['opex']['total']
        ws[f'A{row}'].fill = ExportService.SUBTOTAL_FILL
        ws[f'B{row}'].fill = ExportService.SUBTOTAL_FILL
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2

        # EBITDA
        ws[f'A{row}'] = "RESULTADO OPERATIVO (EBITDA)"
        ws[f'B{row}'] = pnl_data['ebitda']
        ws[f'A{row}'].fill = ExportService.TOTAL_FILL
        ws[f'B{row}'].fill = ExportService.TOTAL_FILL
        ws[f'A{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 1

        ws[f'A{row}'] = "Margen EBITDA %"
        ws[f'B{row}'] = pnl_data['ebitda_margin_pct'] / 100
        ws[f'B{row}'].number_format = '0.00%'

    @staticmethod
    def _write_opex_detail_sheet(ws, pnl_data: dict) -> None:
        """Escribe el desglose detallado de OPEX."""
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18

        row = 1

        # Encabezado
        ws[f'A{row}'] = "Detalle de Gastos Operativos"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 2

        # Tabla
        ws[f'A{row}'] = "CATEGORÍA"
        ws[f'B{row}'] = "MONTO ($)"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.fill = ExportService.HEADER_FILL
            cell.font = ExportService.HEADER_FONT
            cell.border = ExportService.BORDER
        row += 1

        # Datos por categoría
        for category, amount in pnl_data['opex']['by_category'].items():
            ws[f'A{row}'] = category
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1

        # Total
        ws[f'A{row}'] = "TOTAL"
        ws[f'B{row}'] = pnl_data['opex']['total']
        ws[f'A{row}'].fill = ExportService.TOTAL_FILL
        ws[f'B{row}'].fill = ExportService.TOTAL_FILL
        ws[f'A{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].number_format = '#,##0.00'

    @staticmethod
    def _write_cashflow_sheet(ws, cf_data: dict, date_from: date, date_to: date) -> None:
        """Escribe los datos del Cash Flow en la hoja de trabajo."""
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18

        row = 1

        # Encabezado
        ws[f'A{row}'] = "BULONERA ALVEAR"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1

        ws[f'A{row}'] = "Flujo de Caja (Percibido)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'] = f"Período: {date_from} a {date_to}"
        ws[f'A{row}'].font = Font(italic=True, size=10)
        row += 2

        # Encabezado de tabla
        ws[f'A{row}'] = "CONCEPTO"
        ws[f'B{row}'] = "MONTO ($)"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.fill = ExportService.HEADER_FILL
            cell.font = ExportService.HEADER_FONT
            cell.border = ExportService.BORDER
        row += 1

        # COBROS POR MÉTODO
        ws[f'A{row}'] = "COBROS CONFIRMADOS"
        ws[f'A{row}'].fill = ExportService.SECTION_FILL
        ws[f'A{row}'].font = ExportService.SECTION_FONT
        row += 1

        for method, amount in cf_data['inflows']['by_method'].items():
            ws[f'A{row}'] = f"  {method}"
            ws[f'B{row}'] = amount
            ws[f'B{row}'].number_format = '#,##0.00'
            row += 1

        ws[f'A{row}'] = "TOTAL COBROS"
        ws[f'B{row}'] = cf_data['inflows']['total']
        ws[f'A{row}'].fill = ExportService.SUBTOTAL_FILL
        ws[f'B{row}'].fill = ExportService.SUBTOTAL_FILL
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2

        # GASTOS PAGADOS
        ws[f'A{row}'] = "GASTOS PAGADOS"
        ws[f'A{row}'].fill = ExportService.SECTION_FILL
        ws[f'A{row}'].font = ExportService.SECTION_FONT
        row += 1

        ws[f'A{row}'] = "  Total Gastos Pagados"
        ws[f'B{row}'] = -cf_data['outflows']['total']
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2

        # FLUJO NETO
        ws[f'A{row}'] = "FLUJO DE CAJA NETO"
        ws[f'B{row}'] = cf_data['net_cash_flow']
        ws[f'A{row}'].fill = ExportService.TOTAL_FILL
        ws[f'B{row}'].fill = ExportService.TOTAL_FILL
        ws[f'A{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].font = ExportService.TOTAL_FONT
        ws[f'B{row}'].number_format = '#,##0.00'
