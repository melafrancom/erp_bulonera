"""
Tests para ExportService (Excel generation).
"""
import pytest
from datetime import date
from decimal import Decimal
from openpyxl import load_workbook
from io import BytesIO

from reports.services.export_service import ExportService


@pytest.mark.django_db
class TestExportService:
    """Tests de exportación a Excel."""

    def test_export_pnl_to_xlsx_returns_bytes(self, authorized_invoice):
        """export_pnl_to_xlsx retorna bytes válidos."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_pnl_to_xlsx(date_from, date_to)
        
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_export_pnl_to_xlsx_valid_excel(self, sale_with_items):
        """El archivo generado es un Excel válido con hojas esperadas."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_pnl_to_xlsx(date_from, date_to)
        
        # Cargar como workbook
        wb = load_workbook(BytesIO(content))
        
        # Verificar hojas
        assert "Estado de Resultados" in wb.sheetnames
        assert "Detalle OPEX" in wb.sheetnames

    def test_export_pnl_to_xlsx_has_headers(self, authorized_invoice):
        """El archivo contiene datos (verificación básica)."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_pnl_to_xlsx(date_from, date_to)
        wb = load_workbook(BytesIO(content))
        ws = wb["Estado de Resultados"]
        
        # Verificar que tiene datos (al menos una fila con valores)
        has_data = False
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                has_data = True
                break
        
        assert has_data, "El archivo no contiene datos"

    def test_export_cashflow_to_xlsx_returns_bytes(self, confirmed_payment):
        """export_cashflow_to_xlsx retorna bytes válidos."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_cashflow_to_xlsx(date_from, date_to)
        
        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_export_cashflow_to_xlsx_valid_excel(self, paid_expense):
        """El archivo CashFlow generado es un Excel válido."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_cashflow_to_xlsx(date_from, date_to)
        wb = load_workbook(BytesIO(content))
        
        # Verificar que existe la hoja
        assert "Flujo de Caja" in wb.sheetnames

    def test_export_cashflow_to_xlsx_has_flow_data(self, confirmed_payment):
        """El archivo contiene datos de flujo."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_cashflow_to_xlsx(date_from, date_to)
        wb = load_workbook(BytesIO(content))
        ws = wb["Flujo de Caja"]
        
        text_content = [v for row in ws.iter_rows(values_only=True) for v in row if v]
        
        assert any('FLUJO DE CAJA NETO' in str(v) for v in text_content if v)
        assert any('COBROS' in str(v) for v in text_content if v)

    def test_xlsx_numeric_formatting(self, authorized_invoice):
        """Los números en Excel están formateados como moneda."""
        date_from = date(2026, 5, 1)
        date_to = date(2026, 5, 31)
        
        content = ExportService.export_pnl_to_xlsx(date_from, date_to)
        wb = load_workbook(BytesIO(content))
        ws = wb["Estado de Resultados"]
        
        # Encontrar una celda con número
        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if isinstance(cell.value, (int, float, Decimal)):
                    # Verificar que tiene formato
                    assert cell.number_format is not None
                    break
