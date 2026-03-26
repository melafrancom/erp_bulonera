"""
Vistas web (templates) para la app Suppliers.
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
import os
import logging

logger = logging.getLogger(__name__)

from suppliers.models import Supplier, SupplierTag
from suppliers.web.forms import SupplierForm


@login_required
def supplier_list(request):
    """Listado de proveedores con búsqueda y filtros."""
    suppliers = Supplier.objects.prefetch_related('tags').all()

    # Búsqueda
    search = request.GET.get('search', '')
    if search:
        suppliers = suppliers.filter(
            models.Q(business_name__icontains=search) |
            models.Q(trade_name__icontains=search) |
            models.Q(cuit__icontains=search)
        )

    # Filtro por tag
    tag_id = request.GET.get('tag')
    if tag_id:
        suppliers = suppliers.filter(tags__id=tag_id)

    # Filtro por condición IVA
    tax_condition = request.GET.get('tax_condition')
    if tax_condition:
        suppliers = suppliers.filter(tax_condition=tax_condition)

    tags = SupplierTag.objects.all()

    context = {
        'suppliers': suppliers,
        'tags': tags,
        'search': search,
        'selected_tag': tag_id,
        'selected_tax_condition': tax_condition,
    }
    return render(request, 'suppliers/supplier_list.html', context)


@login_required
def supplier_detail(request, pk):
    """Detalle completo de un proveedor."""
    supplier = get_object_or_404(Supplier, pk=pk)
    
    # Manejar eliminación desde la vista de detalle
    if request.method == 'POST' and request.POST.get('action') == 'delete':
        from suppliers.services import SupplierService
        SupplierService.soft_delete(supplier, request.user)
        messages.success(request, f'Proveedor "{supplier.business_name}" eliminado correctamente.')
        return redirect('suppliers_web:supplier_list')

    from products.models import Product
    products = Product.objects.filter(supplier=supplier)

    context = {
        'supplier': supplier,
        'products': products,
    }
    return render(request, 'suppliers/supplier_detail.html', context)


@login_required
def supplier_create(request):
    """Crear proveedor."""
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.created_by = request.user
            supplier.save()
            form.save_m2m()  # Guardar M2M (tags)
            messages.success(request, f'Proveedor "{supplier.business_name}" creado exitosamente.')
            return redirect('suppliers_web:supplier_detail', pk=supplier.pk)
    else:
        form = SupplierForm()

    return render(request, 'suppliers/supplier_form.html', {
        'form': form,
        'title': 'Nuevo Proveedor',
    })


@login_required
def supplier_edit(request, pk):
    """Editar proveedor."""
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            supplier = form.save(commit=False)
            supplier.updated_by = request.user
            supplier.save()
            form.save_m2m()
            messages.success(request, f'Proveedor "{supplier.business_name}" actualizado.')
            return redirect('suppliers_web:supplier_detail', pk=supplier.pk)
    else:
        form = SupplierForm(instance=supplier)

    return render(request, 'suppliers/supplier_form.html', {
        'form': form,
        'supplier': supplier,
        'title': f'Editar Proveedor - {supplier.business_name}',
    })


@login_required
def supplier_import(request):
    """Vista para importar proveedores desde Excel/CSV (GET muestra, POST procesa)."""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "No se seleccionó ningún archivo.")
            return render(request, 'suppliers/supplier_import.html')

        # Validar extensión
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        if ext not in ('.xlsx', '.csv'):
            messages.error(request, f"Formato '{ext}' no soportado. Usá .xlsx o .csv")
            return render(request, 'suppliers/supplier_import.html')

        # Guardar archivo temporal
        imports_dir = os.path.join(settings.MEDIA_ROOT, 'imports')
        os.makedirs(imports_dir, exist_ok=True)

        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"import_suppliers_{timestamp}{ext}"
        file_path = os.path.join(imports_dir, file_name)

        with open(file_path, 'wb') as dest:
            for chunk in uploaded_file.chunks():
                dest.write(chunk)

        # Lanzar tarea Celery si es posible, sino fallback a sync
        try:
            from suppliers.tasks import import_suppliers_task
            task = import_suppliers_task.delay(file_path, request.user.id)
            messages.info(request, "Importación iniciada. Se procesará en segundo plano.")
            # Redirigir a una vista de estado o reporte
            return redirect('suppliers_web:supplier_list') # TODO: Crear vista de reporte si es necesario
        except Exception as e:
            logger.error(f"Error al lanzar tarea de importación de proveedores: {e}", exc_info=True)
            # Fallback sincrónico
            from suppliers.services import SupplierImportService
            service = SupplierImportService()
            report = service.import_from_file(file_path, request.user.id)
            
            if report['status'] == 'completed':
                messages.success(
                    request, 
                    f"Importación completada: {report['created']} creados, {report['updated']} actualizados."
                )
            else:
                messages.error(request, f"Error en importación: {report.get('error', 'Desconocido')}")
            
            return redirect('suppliers_web:supplier_list')

    return render(request, 'suppliers/supplier_import.html')


@login_required
def supplier_download_template(request):
    """Genera y descarga la plantilla Excel para importación de proveedores."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Proveedores'

    # Columnas: (nombre_técnico, descripción_amigable, ancho, es_obligatorio)
    columns = [
        ('business_name', 'Razón Social (obligatorio)', 30, True),
        ('cuit', 'CUIT (XX-XXXXXXXX-X) (obligatorio)', 20, True),
        ('trade_name', 'Nombre Comercial', 25, False),
        ('tax_condition', 'Condición IVA (RI, MONO, EX)', 20, False),
        ('email', 'Email', 25, False),
        ('phone', 'Teléfono', 15, False),
        ('mobile', 'Celular', 15, False),
        ('address', 'Dirección', 30, False),
        ('city', 'Ciudad', 20, False),
        ('state', 'Provincia', 20, False),
        ('zip_code', 'CP', 10, False),
        ('bank_name', 'Banco', 20, False),
        ('cbu', 'CBU', 25, False),
        ('bank_alias', 'Alias', 20, False),
        ('contact_person', 'Contacto', 25, False),
        ('payment_term', 'Plazo Pago (días)', 15, False),
        ('notes', 'Notas', 40, False),
    ]

    # Estilos
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid') # Indigo 600
    
    for col_idx, (col_name, desc, width, req) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = fill
        ws.column_dimensions[cell.column_letter].width = width

    # Ejemplo
    example = {
        'business_name': 'Bulonera Alvear S.A.',
        'cuit': '30-12345678-9',
        'trade_name': 'Bulonera Alvear',
        'tax_condition': 'RI',
        'email': 'ventas@bulonera.com',
        'payment_term': '30',
    }
    for col_idx, (col_name, _, _, _) in enumerate(columns, 1):
        ws.cell(row=2, column=col_idx, value=example.get(col_name, ''))

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_proveedores.xlsx"'
    wb.save(response)
    return response


# Importar models para uso en búsqueda
from django.db import models
