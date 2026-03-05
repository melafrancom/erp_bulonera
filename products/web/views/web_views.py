"""
Vistas web (templates) para la app Products.
"""
import os
import logging
from decimal import Decimal, InvalidOperation

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db.models import Prefetch, Q, Case, When, IntegerField, Value
from django.conf import settings

from products.models import Product, Category, Subcategory, PriceList
from products.services import ProductService, PriceService
from suppliers.models import Supplier

logger = logging.getLogger(__name__)


def _require_product_permission(user):
    """Verifica que el usuario tenga permiso para gestionar productos."""
    return user.can_manage_products if hasattr(user, 'can_manage_products') else user.is_staff


def _parse_product_form(request):
    """
    Extrae y parsea los datos del formulario de producto.
    Retorna dict listo para ProductService.
    """
    POST = request.POST

    data = {
        'code': POST.get('code', '').strip(),
        'name': POST.get('name', '').strip(),
        'description': POST.get('description', '').strip(),
        'brand': POST.get('brand', '').strip(),
        # supplier handled separately via FK
        # Técnicos
        'diameter': POST.get('diameter', '').strip() or None,
        'length': POST.get('length', '').strip() or None,
        'material': POST.get('material', '').strip() or None,
        'grade': POST.get('grade', '').strip() or None,
        'norm': POST.get('norm', '').strip() or None,
        'colour': POST.get('colour', '').strip() or None,
        'product_type': POST.get('product_type', '').strip() or None,
        'form': POST.get('form', '').strip() or None,
        'thread_format': POST.get('thread_format', '').strip() or None,
        'origin': POST.get('origin', '').strip() or None,
        # Stock
        'unit_of_sale': POST.get('unit_of_sale', 'UNIDAD'),
        'stock_control_enabled': POST.get('stock_control_enabled') == 'true',
    }

    # Categoría
    category_id = POST.get('category')
    if category_id:
        try:
            data['category'] = Category.objects.get(id=int(category_id))
        except (Category.DoesNotExist, ValueError):
            raise ValidationError("Categoría seleccionada no válida.")
    else:
        data['category'] = None

    # SKU (opcional)
    sku = POST.get('sku', '').strip()
    if sku:
        data['sku'] = sku

    # Precios — siempre Decimal
    try:
        data['price'] = Decimal(POST.get('price', '0'))
    except (InvalidOperation, TypeError):
        data['price'] = Decimal('0')

    try:
        data['cost'] = Decimal(POST.get('cost', '0'))
    except (InvalidOperation, TypeError):
        data['cost'] = Decimal('0')

    try:
        data['tax_rate'] = Decimal(POST.get('tax_rate', '21.00'))
    except (InvalidOperation, TypeError):
        data['tax_rate'] = Decimal('21.00')

    # Stock
    try:
        data['stock_quantity'] = int(POST.get('stock_quantity', 0))
    except (ValueError, TypeError):
        data['stock_quantity'] = 0

    try:
        data['min_stock'] = int(POST.get('min_stock', 0))
    except (ValueError, TypeError):
        data['min_stock'] = 0

    try:
        data['min_sale_unit'] = int(POST.get('min_sale_unit', 1))
    except (ValueError, TypeError):
        data['min_sale_unit'] = 1

    # Subcategorías (M2M)
    subcat_ids = POST.getlist('subcategories')
    if subcat_ids:
        data['subcategories'] = Subcategory.objects.filter(
            id__in=[int(s) for s in subcat_ids if s.isdigit()]
        )

    # Supplier FK
    supplier_id = POST.get('supplier')
    if supplier_id:
        try:
            data['supplier'] = Supplier.objects.get(id=int(supplier_id))
        except (Supplier.DoesNotExist, ValueError):
            data['supplier'] = None
    else:
        data['supplier'] = None

    return data


# =============================================================================
# LISTADO
# =============================================================================

@login_required
def product_list(request):
    """Listado de productos con búsqueda y filtros."""
    queryset = Product.objects.select_related(
        'category'
    ).prefetch_related('subcategories').all()

    # Búsqueda general por nombre, código, sku, o descripción
    search = request.GET.get('search', '').strip()
    if search:
        words = [w for w in search.split() if w]
        
        # Coincidencia por código, sku o descripción con el término completo
        exact_q = (
            Q(code__icontains=search) |
            Q(sku__icontains=search) |
            Q(description__icontains=search)
        )
        
        # Coincidencia por nombre (debe contener TODAS las palabras sin importar el orden)
        name_q = Q()
        if words:
            name_q = Q(name__icontains=words[0])
            for w in words[1:]:
                name_q &= Q(name__icontains=w)

        # Filtramos el listado
        queryset = queryset.filter(exact_q | name_q).distinct()
        
        # Anotamos match_score=1 para mantener la compatibilidad con el template y resaltarlos
        queryset = queryset.annotate(
            match_score=Value(1, output_field=IntegerField())
        ).order_by('name')
    else:
        # Sin búsqueda
        queryset = queryset.annotate(
            match_score=Value(0, output_field=IntegerField())
        ).order_by('name')

    # Filtros
    code_filter = request.GET.get('code', '').strip()
    if code_filter:
        queryset = queryset.filter(code__icontains=code_filter)

    sku_filter = request.GET.get('sku', '').strip()
    if sku_filter:
        queryset = queryset.filter(sku__icontains=sku_filter)

    brand_filter = request.GET.get('brand', '').strip()
    if brand_filter:
        queryset = queryset.filter(brand__icontains=brand_filter)

    supplier_filter = request.GET.get('supplier', '').strip()
    if supplier_filter:
        queryset = queryset.filter(
            Q(supplier__business_name__icontains=supplier_filter) |
            Q(supplier__trade_name__icontains=supplier_filter)
        )

    category_filter = request.GET.get('category', '').strip()
    if category_filter:
        queryset = queryset.filter(category_id=category_filter)

    subcategory_filter = request.GET.get('subcategory', '').strip()
    if subcategory_filter:
        queryset = queryset.filter(subcategories__id=subcategory_filter).distinct()

    # Paginación
    paginator = Paginator(queryset, 50)
    page_number = request.GET.get('page')
    products = paginator.get_page(page_number)

    # Obtener todas las listas de precios activas
    price_lists = PriceList.objects.filter(is_active=True).order_by('priority')

    # Computar Listas de Precio SOLAMENTE para los productos de la página actual
    from decimal import Decimal, ROUND_HALF_UP
    for p in products:
        # Calcular IVA del precio base
        tax_multiplier = 1 + (p.tax_rate / Decimal('100.0'))
        p_base_with_tax = (p.price * tax_multiplier).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        p.base_price_with_tax = p_base_with_tax
        
        # p.price es Decimal, tax_rate es Decimal
        p.computed_prices = []
        for pl in price_lists:
            calc = pl.calculate_price(p.price, p.tax_rate)
            p.computed_prices.append({
                'list_name': pl.name,
                'type': pl.list_type,
                'percentage': pl.percentage,
                'price_without_tax': calc['price_without_tax'],
                'price_with_tax': calc['price_with_tax'],
            })

    # Datos para filtros
    categories = Category.objects.all().order_by('name')
    subcategories = Subcategory.objects.all().order_by('name')

    context = {
        'products': products,
        'categories': categories,
        'subcategories': subcategories,
        'search': search,
        'code_filter': code_filter,
        'sku_filter': sku_filter,
        'brand_filter': brand_filter,
        'supplier_filter': supplier_filter,
        'category_filter': category_filter,
        'subcategory_filter': subcategory_filter,
        'total_count': paginator.count,
    }
    return render(request, 'products/product_list.html', context)


# =============================================================================
# DETALLE
# =============================================================================

@login_required
def product_detail(request, pk):
    """Detalle de producto con listas de precios calculadas."""
    product = get_object_or_404(
        Product.objects.select_related('category').prefetch_related(
            'subcategories', 'images'
        ),
        pk=pk
    )

    # Calcular precios con listas
    price_service = PriceService()
    price_data = price_service.calculate_prices_with_lists(product)

    context = {
        'product': product,
        'price_data': price_data,
    }
    return render(request, 'products/product_detail.html', context)


# =============================================================================
# CREAR
# =============================================================================

@login_required
def product_create(request):
    """Formulario de creación de producto (GET muestra, POST crea)."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permiso para crear productos.")
        return redirect('products:product_list')

    if request.method == 'POST':
        try:
            data = _parse_product_form(request)
            service = ProductService()
            product = service.create_product(data, request.user)
            messages.success(
                request,
                f"Producto '{product.code} - {product.name}' creado exitosamente."
            )
            return redirect('products:product_detail', pk=product.pk)
        except ValidationError as e:
            error_msg = e.message if hasattr(e, 'message') else str(e)
            messages.error(request, f"Error al crear producto: {error_msg}")
        except Exception as e:
            logger.error(f"Error inesperado al crear producto: {e}", exc_info=True)
            messages.error(request, f"Error inesperado: {e}")

    categories = Category.objects.all().order_by('name')
    subcategories = Subcategory.objects.all().order_by('name')
    suppliers = Supplier.objects.all().order_by('business_name')

    context = {
        'categories': categories,
        'subcategories': subcategories,
        'suppliers': suppliers,
        'unit_choices': Product.UNIT_CHOICES,
        'condition_choices': Product.CONDITION_CHOICES,
        'is_edit': False,
    }
    return render(request, 'products/product_form.html', context)


# =============================================================================
# EDITAR
# =============================================================================

@login_required
def product_edit(request, pk):
    """Formulario de edición de producto (GET muestra, POST actualiza)."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permiso para editar productos.")
        return redirect('products:product_list')

    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        try:
            data = _parse_product_form(request)
            service = ProductService()
            service.update_product(product, data, request.user)
            messages.success(
                request,
                f"Producto '{product.code}' actualizado exitosamente."
            )
            return redirect('products:product_detail', pk=product.pk)
        except ValidationError as e:
            error_msg = e.message if hasattr(e, 'message') else str(e)
            messages.error(request, f"Error al actualizar: {error_msg}")
        except Exception as e:
            logger.error(f"Error inesperado al editar producto: {e}", exc_info=True)
            messages.error(request, f"Error inesperado: {e}")

    categories = Category.objects.all().order_by('name')
    subcategories = Subcategory.objects.all().order_by('name')
    suppliers = Supplier.objects.all().order_by('business_name')

    context = {
        'product': product,
        'categories': categories,
        'subcategories': subcategories,
        'suppliers': suppliers,
        'unit_choices': Product.UNIT_CHOICES,
        'condition_choices': Product.CONDITION_CHOICES,
        'is_edit': True,
    }
    return render(request, 'products/product_form.html', context)


# =============================================================================
# ELIMINAR (soft delete)
# =============================================================================

@login_required
def product_delete(request, pk):
    """Soft delete de producto (solo POST)."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permiso para eliminar productos.")
        return redirect('products:product_list')

    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        code = product.code
        name = product.name
        product.delete(user=request.user)
        messages.success(request, f"Producto '{code} - {name}' eliminado.")
        return redirect('products:product_list')

    # Si llega por GET, redirigir al detalle
    return redirect('products:product_detail', pk=pk)


# =============================================================================
# IMPORTAR
# =============================================================================

@login_required
def product_import(request):
    """Vista para importar productos desde Excel (GET muestra, POST procesa)."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permiso para importar productos.")
        return redirect('products:product_list')

    if request.method == 'POST':
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            messages.error(request, "No se seleccionó ningún archivo.")
            return render(request, 'products/import_products.html')

        # Validar extensión
        ext = os.path.splitext(uploaded_file.name)[1].lower()
        if ext not in ('.xlsx', '.csv'):
            messages.error(request, f"Formato '{ext}' no soportado. Usá .xlsx o .csv")
            return render(request, 'products/import_products.html')

        # Guardar archivo temporal
        imports_dir = os.path.join(settings.MEDIA_ROOT, 'imports')
        os.makedirs(imports_dir, exist_ok=True)

        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        file_name = f"import_{timestamp}{ext}"
        file_path = os.path.join(imports_dir, file_name)

        with open(file_path, 'wb') as dest:
            for chunk in uploaded_file.chunks():
                dest.write(chunk)

        # Lanzar tarea Celery
        try:
            from products.tasks.import_tasks import import_products_from_excel
            task = import_products_from_excel.delay(file_path, request.user.id)
            messages.info(request, "Importación iniciada. Se procesará en segundo plano.")
            return redirect('products:import_report_task', task_id=task.id)
        except Exception as e:
            logger.error(f"Error al lanzar tarea de importación: {e}", exc_info=True)
            messages.error(request, f"Error al iniciar importación: {e}")

    return render(request, 'products/import_products.html')


# =============================================================================
# REPORTE DE IMPORTACIÓN
# =============================================================================

@login_required
def import_report(request, task_id=None):
    """Vista para mostrar resultados de importación."""
    task_status = None
    task_result = None

    if task_id:
        try:
            from celery.result import AsyncResult
            result = AsyncResult(task_id)
            task_status = result.status
            if result.ready():
                task_result = result.result
            elif result.status == 'PROGRESS':
                task_result = result.info
        except Exception as e:
            logger.error(f"Error al consultar tarea {task_id}: {e}")
            task_status = 'UNKNOWN'

    context = {
        'task_id': task_id,
        'task_status': task_status,
        'task_result': task_result,
    }
    return render(request, 'products/import_report.html', context)


# =============================================================================
# DESCARGAR PLANTILLA DE IMPORTACIÓN
# =============================================================================

@login_required
def download_import_template(request):
    """Genera y descarga la plantilla Excel para importación de productos."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.comments import Comment
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Productos'

    # Columnas: (nombre_técnico, descripción_amigable, ancho, es_obligatorio)
    columns = [
        ('code', 'Código del producto (obligatorio)', 15, True),
        ('price', 'Precio de venta (obligatorio)', 12, True),
        ('cost', 'Costo sin IVA (Opcional)', 12, False),
        ('name', 'Nombre del producto', 35, False),
        ('diameter', 'Diámetro (ej: M10)', 12, False),
        ('length', 'Largo (ej: 50mm)', 12, False),
        ('description', 'Descripción del producto', 40, False),
        ('supplier', 'Proveedor (Nombre)', 25, False),
        ('supplier_cuit', 'CUIT del Proveedor (Opcional)', 20, False),
        ('images', 'URLs de imágenes separadas por coma', 30, False),
        ('stock', 'Cantidad en stock', 10, False),
        ('category', 'Nombre de categoría (se crea si no existe)', 20, False),
        ('subcategories', 'Subcategorías separadas por coma', 25, False),
        ('brand', 'Marca del producto', 18, False),
        ('condition', 'Estado: new, used, refurbished', 15, False),
        ('gallery', 'URLs de galería separadas por coma', 30, False),
        ('norm', 'Norma técnica (ej: DIN 933)', 15, False),
        ('grade', 'Grado (ej: 8.8)', 12, False),
        ('material', 'Material (ej: Acero)', 18, False),
        ('colour', 'Color', 12, False),
        ('type', 'Tipo de producto', 15, False),
        ('form', 'Forma', 15, False),
        ('thread_formats', 'Formato de rosca', 18, False),
        ('origin', 'País de origen', 15, False),
        ('faq', 'Preguntas frecuentes', 30, False),
        ('gtin', 'Código GTIN/EAN', 18, False),
        ('mpn', 'Número de parte del fabricante', 18, False),
        ('meta_title', 'Título SEO', 25, False),
        ('meta_description', 'Descripción SEO', 35, False),
        ('meta_keywords', 'Palabras clave SEO', 25, False),
        ('google_category', 'Categoría de Google Shopping', 25, False),
    ]

    # Estilos
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    required_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    optional_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB'),
    )

    # Fila 1: nombres técnicos (pandas lee esta fila como headers)
    for col_idx, (col_name, description, width, required) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = required_fill if required else optional_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.comment = Comment(description, 'ERP Bulonera')
        ws.column_dimensions[cell.column_letter].width = width

    # Fila 2: ejemplo de datos (en itálica gris)
    example_font = Font(name='Calibri', size=10, italic=True, color='6B7280')
    example_data = {
        'code': 'BUL-001',
        'price': '1250.50',
        'cost': '950.00',
        'name': 'Arandela Plana',
        'diameter': 'M10',
        'length': '50mm',
        'description': 'Arandela plana de acero',
        'supplier': 'Bulonera Alvear S.A.',
        'supplier_cuit': '30-11111111-2',
        'images': 'arandelaPlana.png',
        'stock': '100',
        'category': 'Bulonería',
        'subcategories': 'Arandelas',
        'brand': 'Stanley',
        'condition': 'new',
        'gallery': 'arandelaPlana.png, arandelaPlana2.png',
        'norm': 'DIN 933',
        'grade': '8.8',
        'material': 'Acero',
        'colour': 'Gris',
        'type': 'Tipo',
        'form': 'Forma',
        'thread_formats': 'Formato de rosca',
        'origin': 'Nacional',
        'faq': 'Preguntas frecuentes',
        'gtin': 'Código GTIN/EAN',
        'mpn': 'Número de parte del fabricante',
        'meta_title': 'Título SEO',
        'meta_description': 'Descripción SEO',
        'meta_keywords': 'Palabras clave SEO',
        'google_category': 'Bricolaje > Accesorios de bricolaje > Artículos de ferretería > Arandelas',
    }
    for col_idx, (col_name, _, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=example_data.get(col_name, ''))
        cell.font = example_font
        cell.border = thin_border

    # Congelar panel debajo del header
    ws.freeze_panes = 'A2'
    # Altura de fila del header
    ws.row_dimensions[1].height = 25

    # Generar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="plantilla_importacion_productos.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# LISTAS DE PRECIOS — CRUD
# =============================================================================

@login_required
def pricelist_list(request):
    """Listado de listas de precios con búsqueda y paginación."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permisos para gestionar listas de precios.")
        return redirect('core:dashboard')

    pricelists = PriceList.objects.filter(is_active=True)

    search = request.GET.get('search', '').strip()
    if search:
        pricelists = pricelists.filter(name__icontains=search)

    pricelists = pricelists.order_by('priority', 'name')

    paginator = Paginator(pricelists, 25)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    context = {
        'page_obj': page_obj,
        'pricelists': page_obj.object_list,
        'search': search,
        'total_count': paginator.count,
    }
    return render(request, 'products/pricelist_list.html', context)


@login_required
def pricelist_create(request):
    """Crear nueva lista de precios."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permisos para crear listas de precios.")
        return redirect('products:pricelist_list')

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            list_type = request.POST.get('list_type', '')
            percentage = request.POST.get('percentage', '')
            description = request.POST.get('description', '').strip()
            priority = request.POST.get('priority', '0')

            if not name:
                raise ValueError("El nombre es obligatorio.")
            if list_type not in ('DISCOUNT', 'SURCHARGE'):
                raise ValueError("Tipo de lista inválido.")
            if not percentage:
                raise ValueError("El porcentaje es obligatorio.")

            percentage = Decimal(percentage)
            priority = int(priority) if priority else 0

            if percentage <= 0:
                raise ValueError("El porcentaje debe ser mayor a 0.")

            pricelist = PriceList.objects.create(
                name=name,
                list_type=list_type,
                percentage=percentage,
                description=description or '',
                priority=priority,
                created_by=request.user,
                updated_by=request.user,
            )
            messages.success(request, f'Lista de precios "{pricelist.name}" creada exitosamente.')
            return redirect('products:pricelist_list')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f"Error al crear lista de precios: {e}", exc_info=True)
            messages.error(request, f'Error al crear la lista: {e}')

    return render(request, 'products/pricelist_form.html', {'is_edit': False})


@login_required
def pricelist_edit(request, pk):
    """Editar lista de precios existente."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permisos para editar listas de precios.")
        return redirect('products:pricelist_list')

    pricelist = get_object_or_404(PriceList, pk=pk, is_active=True)

    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            list_type = request.POST.get('list_type', '')
            percentage = request.POST.get('percentage', '')
            description = request.POST.get('description', '').strip()
            priority = request.POST.get('priority', '0')

            if not name:
                raise ValueError("El nombre es obligatorio.")
            if list_type not in ('DISCOUNT', 'SURCHARGE'):
                raise ValueError("Tipo de lista inválido.")
            if not percentage:
                raise ValueError("El porcentaje es obligatorio.")

            percentage = Decimal(percentage)
            priority = int(priority) if priority else 0

            if percentage <= 0:
                raise ValueError("El porcentaje debe ser mayor a 0.")

            pricelist.name = name
            pricelist.list_type = list_type
            pricelist.percentage = percentage
            pricelist.description = description or ''
            pricelist.priority = priority
            pricelist.updated_by = request.user
            pricelist.save()

            messages.success(request, f'Lista de precios "{pricelist.name}" actualizada exitosamente.')
            return redirect('products:pricelist_list')

        except (ValueError, InvalidOperation) as e:
            messages.error(request, str(e))
        except Exception as e:
            logger.error(f"Error al actualizar lista de precios: {e}", exc_info=True)
            messages.error(request, f'Error al actualizar la lista: {e}')

    return render(request, 'products/pricelist_form.html', {
        'is_edit': True,
        'pricelist': pricelist,
    })


@login_required
def pricelist_delete(request, pk):
    """Soft delete de lista de precios (solo POST)."""
    if not _require_product_permission(request.user):
        messages.error(request, "No tenés permisos para eliminar listas de precios.")
        return redirect('products:pricelist_list')

    if request.method != 'POST':
        return redirect('products:pricelist_list')

    pricelist = get_object_or_404(PriceList, pk=pk, is_active=True)

    try:
        pricelist.delete(user=request.user)
        messages.success(request, f'Lista de precios "{pricelist.name}" eliminada exitosamente.')
    except Exception as e:
        logger.error(f"Error al eliminar lista de precios: {e}", exc_info=True)
        messages.error(request, f'Error al eliminar la lista: {e}')

    return redirect('products:pricelist_list')
