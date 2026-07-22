"""
Servicios de negocio para la app Products.

Contiene la lógica de negocio separada de las vistas:
- ProductService: CRUD y operaciones sobre productos
- PriceService: Cálculos de precios con listas
- ProductImportService: Importación masiva desde Excel/CSV
- ProductExportService: Exportación a Excel
"""

import os
import logging
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.core.exceptions import ValidationError
from django.conf import settings

import pandas as pd

from products.models import Product, Category, Subcategory, PriceList

logger = logging.getLogger('api')


# =============================================================================
# ProductService
# =============================================================================

class ProductService:
    """Lógica de negocio de productos."""

    @transaction.atomic
    def create_product(self, data, user):
        """
        Crea un producto nuevo.

        Args:
            data: dict con campos del producto
            user: usuario que crea

        Returns:
            Product creado

        Raises:
            ValidationError si el código ya existe o datos inválidos
        """
        code = data.get('code', '').strip()
        if not code:
            raise ValidationError("El código del producto es obligatorio.")

        if Product.all_objects.filter(code=code).exists():
            raise ValidationError(f"El código '{code}' ya existe.")

        # Extraer M2M
        subcategories_data = data.pop('subcategories', [])

        product = Product(**data)
        product.created_by = user
        product.full_clean()
        product.save()

        if subcategories_data:
            product.subcategories.set(subcategories_data)

        logger.info(
            f"Producto creado: {product.code} - {product.name} "
            f"por {user.username}"
        )
        return product

    @transaction.atomic
    def update_product(self, product, data, user):
        """Actualiza un producto existente."""
        subcategories_data = data.pop('subcategories', None)

        for field, value in data.items():
            setattr(product, field, value)

        product.updated_by = user
        product.full_clean()
        product.save()

        if subcategories_data is not None:
            product.subcategories.set(subcategories_data)

        logger.info(
            f"Producto actualizado: {product.code} por {user.username}"
        )
        return product

    @transaction.atomic
    def update_price(self, product_id, price_data, user):
        """
        Actualización rápida de precios.

        Args:
            product_id: ID del producto
            price_data: dict con sale_price y/o cost_price
            user: usuario que actualiza

        Returns:
            Product actualizado
        """
        product = Product.objects.select_for_update().get(id=product_id)

        if 'sale_price' in price_data:
            new_price = Decimal(str(price_data['sale_price']))
            if new_price < 0:
                raise ValidationError("El precio de venta no puede ser negativo.")
            product.price = new_price

        if 'cost_price' in price_data:
            new_cost = Decimal(str(price_data['cost_price']))
            if new_cost < 0:
                raise ValidationError("El precio de costo no puede ser negativo.")
            product.cost = new_cost

        if 'tax_rate' in price_data:
            product.tax_rate = Decimal(str(price_data['tax_rate']))

        product.updated_by = user
        product.save()

        logger.info(
            f"Precio actualizado: {product.code} → "
            f"venta={product.price}, costo={product.cost} "
            f"por {user.username}"
        )
        return product

    @transaction.atomic
    def soft_delete(self, product, user):
        """Soft delete de un producto."""
        product.delete(user=user)
        logger.info(
            f"Producto eliminado (soft): {product.code} por {user.username}"
        )

    @staticmethod
    def generate_barcode_image(barcode_text: str) -> bytes:
        """Genera una imagen de código de barras (Code128 o EAN13) en memoria y retorna los bytes PNG."""
        import io
        import barcode
        from barcode.writer import ImageWriter

        text_clean = str(barcode_text).strip()
        if not text_clean:
            raise ValidationError("El texto del código de barras no puede estar vacío.")

        try:
            if len(text_clean) == 13 and text_clean.isdigit():
                coder = barcode.get_barcode_class('ean13')
            else:
                coder = barcode.get_barcode_class('code128')

            rv = io.BytesIO()
            # Escribir el código de barras en formato PNG
            coder(text_clean, writer=ImageWriter()).write(rv)
            return rv.getvalue()
        except Exception as e:
            logger.exception(f"Error generando código de barras para '{barcode_text}': {e}")
            raise ValidationError(f"No se pudo generar el código de barras: {e}")

    @staticmethod
    def generate_qr_image(qr_text: str) -> bytes:
        """Genera una imagen de código QR en memoria a partir de un texto y retorna los bytes PNG."""
        import io
        import qrcode

        text_clean = str(qr_text).strip()
        if not text_clean:
            raise ValidationError("El texto del código QR no puede estar vacío.")

        try:
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(text_clean)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")

            rv = io.BytesIO()
            img.save(rv, format="PNG")
            return rv.getvalue()
        except Exception as e:
            logger.exception(f"Error generando código QR para '{qr_text}': {e}")
            raise ValidationError(f"No se pudo generar el código QR: {e}")


# =============================================================================
# PriceService
# =============================================================================

class PriceService:
    """Cálculos de precios con listas de bonificación/recargo."""

    def calculate_prices_with_lists(self, product):
        """
        Calcula precios del producto con todas las listas activas.

        Returns:
            dict con base_price y price_lists
        """
        base_price = {
            'sale_price_without_tax': product.price,
            'sale_price_with_tax': product.sale_price_with_tax,
            'cost_price': product.cost,
            'tax_rate': product.tax_rate,
            'profit_margin_percentage': product.profit_margin_percentage,
            'profit_amount': product.profit_amount,
        }

        price_lists = []
        for plist in PriceList.objects.filter(is_active=True).order_by('priority'):
            prices = plist.calculate_price(product.price, product.tax_rate)
            price_lists.append({
                'id': plist.id,
                'name': plist.name,
                'type': plist.list_type,
                'percentage': plist.percentage,
                **prices,
            })

        return {
            'base_price': base_price,
            'price_lists': price_lists,
        }


# =============================================================================
# ProductImportService
# =============================================================================

class ProductImportService:
    """Importación masiva de productos desde archivos Excel/CSV."""

    REQUIRED_COLUMNS = ['code', 'price']
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
    VALID_EXTENSIONS = ['.xlsx', '.csv']

    def validate_file(self, file):
        """Valida el archivo antes de procesarlo."""
        ext = os.path.splitext(file.name)[1].lower()
        if ext not in self.VALID_EXTENSIONS:
            raise ValidationError(
                f"Formato no soportado: {ext}. Use .xlsx o .csv"
            )
        if file.size > self.MAX_FILE_SIZE:
            raise ValidationError(
                f"Archivo demasiado grande ({file.size / 1024 / 1024:.1f} MB). "
                f"Máximo: {self.MAX_FILE_SIZE / 1024 / 1024:.0f} MB."
            )

    def import_from_file(self, file_path, user_id, update_state_callback=None):
        """
        Importa productos desde un archivo Excel/CSV.

        Args:
            file_path: ruta absoluta del archivo
            user_id: ID del usuario que importa
            update_state_callback: función para actualizar progreso (Celery)

        Returns:
            dict con reporte de importación
        """
        from core.models import User
        user = User.objects.get(id=user_id)

        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.csv':
            df = pd.read_csv(file_path, keep_default_na=True, na_values=[''])
        else:
            df = pd.read_excel(file_path, keep_default_na=True, na_values=[''])

        df = df.where(pd.notna(df), None)

        total_rows = len(df)
        successful = 0
        failed = 0
        created = 0
        updated = 0
        errors = []

        for index, row in df.iterrows():
            row_dict = row.to_dict()
            try:
                result = self._process_row(row_dict, user)
                successful += 1
                if result == 'created':
                    created += 1
                else:
                    updated += 1
            except Exception as e:
                failed += 1
                errors.append({
                    'row': index + 2,  # +2: Excel 1-indexed + header
                    'code': str(row_dict.get('code', 'N/A')),
                    'error': str(e),
                })

            # Actualizar progreso cada 10 filas
            if update_state_callback and (index + 1) % 10 == 0:
                update_state_callback({
                    'progress': int((index + 1) / total_rows * 100),
                    'total_rows': total_rows,
                    'processed_rows': index + 1,
                })

        return {
            'total_rows': total_rows,
            'successful': successful,
            'failed': failed,
            'created': created,
            'updated': updated,
            'errors': errors,
        }

    def _resolve_category(self, row, user, is_new, product):
        """Resuelve o crea la categoría asignada de forma normalizada."""
        cat_name = row.get('category')
        if cat_name and not (isinstance(cat_name, float) and pd.isna(cat_name)):
            cat_name = ' '.join(str(cat_name).strip().split())
            category, _ = Category.objects.get_or_create(
                name=cat_name,
                defaults={'created_by': user}
            )
            product.category = category
        elif is_new and not product.category_id:
            category, _ = Category.objects.get_or_create(
                name="Sin categoría",
                defaults={'created_by': user}
            )
            product.category = category

    def _resolve_supplier(self, row, user, product):
        """Resuelve o crea el proveedor asignado al producto."""
        from suppliers.models import Supplier
        sup_name = row.get('supplier')
        sup_cuit = row.get('supplier_cuit')
        if sup_name and not (isinstance(sup_name, float) and pd.isna(sup_name)):
            sup_name = ' '.join(str(sup_name).strip().split())
            sup_cuit = str(sup_cuit).strip() if sup_cuit and not (isinstance(sup_cuit, float) and pd.isna(sup_cuit)) else None
            
            supplier = None
            if sup_cuit:
                supplier = Supplier.objects.filter(cuit=sup_cuit).first()
            if not supplier:
                supplier = Supplier.objects.filter(business_name__iexact=sup_name).first()
                
            if not supplier:
                supplier = Supplier.objects.create(
                    business_name=sup_name,
                    cuit=sup_cuit,
                    created_by=user,
                )
            product.supplier = supplier

    def _process_main_image(self, row, product):
        """Asigna la imagen principal y emite una advertencia si el archivo no existe en el almacenamiento."""
        from django.core.files.storage import default_storage
        img_val = row.get('images') or row.get('image')
        if img_val and not (isinstance(img_val, float) and pd.isna(img_val)):
            first_image = str(img_val).split(',')[0].strip()
            if first_image:
                expected_path = f"photos/products/original/{first_image}"
                product.main_image = expected_path
                if not default_storage.exists(expected_path):
                    logger.warning(
                        f"Imagen principal '{expected_path}' asignada para producto {product.code}, pero no existe en storage."
                    )

    def _process_subcategories_and_gallery(self, row, product, user):
        """Procesa relaciones M2M de subcategorías y la galería de imágenes."""
        subcat_val = row.get('subcategories')
        if subcat_val and not (isinstance(subcat_val, float) and pd.isna(subcat_val)):
            subcat_names = [' '.join(s.strip().split()) for s in str(subcat_val).split(',') if s.strip()]
            subcats = []
            for sname in subcat_names:
                subcat, _ = Subcategory.objects.get_or_create(
                    name=sname,
                    defaults={
                        'category': product.category,
                        'created_by': user,
                    }
                )
                subcats.append(subcat)
            product.subcategories.set(subcats)

        if hasattr(product, '_pending_gallery'):
            from products.models import ProductImage
            product.images.all().delete()
            for idx, img_name in enumerate(product._pending_gallery):
                expected_path = f"photos/products/original/{img_name}"
                ProductImage.objects.create(
                    product=product,
                    image=expected_path,
                    order=idx
                )

    @transaction.atomic
    def _process_row(self, row, user):
        """Procesa una fila del archivo y crea/actualiza un producto."""
        # Validar campos obligatorios
        code = row.get('code')
        price_val = row.get('price')

        if not code or (isinstance(code, float) and pd.isna(code)):
            raise ValueError("Campo 'code' vacío o ausente.")
        if not price_val and price_val != 0:
            raise ValueError("Campo 'price' vacío o ausente.")

        code = str(code).strip()

        try:
            sale_price = Decimal(str(price_val))
        except (InvalidOperation, TypeError):
            raise ValueError(f"Precio inválido: '{price_val}'")

        if sale_price < 0:
            raise ValueError(f"Precio negativo: {sale_price}")

        # Buscar producto existente
        try:
            product = Product.all_objects.get(code=code)
            is_new = False
        except Product.DoesNotExist:
            product = Product(code=code, created_by=user)
            is_new = True

        # Precio
        product.price = sale_price
        product.updated_by = user

        # Costo (M-05: Loguear si es inválido)
        cost_val = row.get('cost_price') or row.get('cost')
        if cost_val and not (isinstance(cost_val, float) and pd.isna(cost_val)):
            try:
                product.cost = Decimal(str(cost_val))
            except (InvalidOperation, TypeError):
                logger.warning(f"Costo inválido ignorado '{cost_val}' en producto {code}")

        # Nombre
        name = row.get('name')
        if name and not (isinstance(name, float) and pd.isna(name)):
            product.name = str(name).strip()
        elif is_new:
            product.name = f"Producto {code}"

        # Categoría y Proveedor
        self._resolve_category(row, user, is_new, product)
        self._resolve_supplier(row, user, product)

        # Campos opcionales simples
        OPTIONAL_TEXT_FIELDS = {
            'sku': 'sku', 'diameter': 'diameter', 'length': 'length',
            'brand': 'brand',
            'material': 'material', 'grade': 'grade', 'norm': 'norm',
            'colour': 'colour', 'type': 'product_type', 'form': 'form',
            'thread_formats': 'thread_format', 'thread_format': 'thread_format',
            'origin': 'origin', 'barcode': 'barcode',
            'gtin': 'gtin', 'mpn': 'mpn',
            'meta_title': 'meta_title', 'meta_description': 'meta_description',
            'meta_keywords': 'meta_keywords', 'google_category': 'google_category',
            'condition': 'condition',
            'other codes': 'other_codes',
            'other_codes': 'other_codes',
            'description': 'description',
            'descripción': 'description',
            'descripcion': 'description',
            'detalle': 'description',
        }
        for csv_col, model_field in OPTIONAL_TEXT_FIELDS.items():
            val = row.get(csv_col)
            if val and not (isinstance(val, float) and pd.isna(val)):
                setattr(product, model_field, str(val).strip())

        # Normalizar condition a lowercase
        if product.condition:
            product.condition = product.condition.lower()
            valid_conditions = [c[0] for c in Product.CONDITION_CHOICES]
            if product.condition not in valid_conditions:
                product.condition = 'new'

        # Stock
        target_stock = None
        stock_val = row.get('stock') or row.get('stock_quantity')
        if stock_val and not (isinstance(stock_val, float) and pd.isna(stock_val)):
            try:
                target_stock = int(float(stock_val))
            except (ValueError, TypeError):
                pass

        # IVA
        tax_val = row.get('tax_rate')
        if tax_val and not (isinstance(tax_val, float) and pd.isna(tax_val)):
            try:
                tax_dec = Decimal(str(tax_val))
                from products.models import validate_afip_tax_rate
                validate_afip_tax_rate(tax_dec)
                product.tax_rate = tax_dec
            except (InvalidOperation, TypeError, ValidationError) as e:
                logger.warning(f"Alícuota IVA inválida '{tax_val}' en producto {code}: {e}")

        # Imagen Principal
        self._process_main_image(row, product)

        # Galería de imágenes (pendiente para después del save)
        gallery_val = row.get('gallery')
        if gallery_val and not (isinstance(gallery_val, float) and pd.isna(gallery_val)):
            gallery_images = [img.strip() for img in str(gallery_val).split(',') if img.strip()]
            if gallery_images:
                setattr(product, '_pending_gallery', gallery_images)

        # Si el producto fue softdeleted, restaurarlo
        if not is_new and not product.is_active:
            product.is_active = True
            product.deleted_at = None
            product.deleted_by = None

        previous_stock = product.stock_quantity if not is_new else 0
        product.save()

        # Ajuste de stock mediante InventoryService si fue provisto
        if target_stock is not None:
            if is_new or previous_stock != target_stock:
                try:
                    from inventory.services import InventoryService
                    inv_service = InventoryService()
                    inv_service.adjust_stock(
                        product_id=product.id,
                        new_quantity=target_stock,
                        reason="Importación masiva de productos",
                        user=user
                    )
                except Exception:
                    if product.stock_quantity != target_stock:
                        product.stock_quantity = target_stock
                        product.save(update_fields=['stock_quantity'])

        # Subcategorías y Galería (post-save)
        self._process_subcategories_and_gallery(row, product, user)

        return 'created' if is_new else 'updated'


# =============================================================================
# ProductExportService
# =============================================================================

class ProductExportService:
    """Exportación de productos a Excel."""

    def export_to_excel(self, queryset=None, file_path=None):
        """
        Exporta productos a un archivo Excel interno (ERP).

        Args:
            queryset: QuerySet de productos (default: todos activos)
            file_path: ruta de destino (default: media/exports/)

        Returns:
            ruta absoluta del archivo generado
        """
        if queryset is None:
            queryset = Product.objects.select_related('category').all()

        data = []
        for p in queryset:
            data.append({
                'code': p.code,
                'sku': p.sku,
                'name': p.name,
                'category': getattr(p.category, 'name', '') if p.category else '',
                'subcategories': ', '.join(
                    p.subcategories.values_list('name', flat=True)
                ),
                'price': float(p.price),
                'cost': float(p.cost),
                'tax_rate': float(p.tax_rate),
                'stock': p.stock_quantity,
                'brand': p.brand,
                'supplier': getattr(p.supplier, 'business_name', '') if p.supplier else '',
                'other_codes': p.other_codes or '',
                'diameter': p.diameter or '',
                'length': p.length or '',
                'material': p.material or '',
                'grade': p.grade or '',
                'norm': p.norm or '',
                'colour': p.colour or '',
                'form': p.form or '',
                'thread_format': p.thread_format or '',
                'origin': p.origin or '',
                'condition': p.condition,
                'unit_of_sale': p.unit_of_sale,
            })

        import io
        df = pd.DataFrame(data)

        if file_path is None:
            bio = io.BytesIO()
            df.to_excel(bio, index=False, engine='openpyxl')
            bio.seek(0)
            logger.info("Productos exportados a memoria (BytesIO)")
            return bio

        df.to_excel(file_path, index=False, engine='openpyxl')
        logger.info(f"Productos exportados a: {file_path}")
        return file_path

    def export_for_web(self, queryset=None, file_path=None):
        """
        Exporta productos a Excel en el formato exacto requerido por la web
        de Bulonera Alvear, listo para importar en la app web.

        Columnas exportadas (siguiendo la política de importación web):
        ▸ OBLIGATORIAS: code, price
        ▸ OPCIONALES: category, name, diameter, length, description, stock,
          subcategories, images, gallery, brand, condition, gtin, mpn
        ▸ ESPECIFICACIONES: norm, grade, material, colour, type, form,
          thread_formats, origin
        ▸ SEO: meta_title, meta_description, meta_keywords, google_category

        Notas:
        - El nombre se exporta como nombre BASE (sin dimensiones) porque
          la web los concatena automáticamente con diameter + length.
        - stock se exporta como entero.
        - La imagen principal se exporta como ruta relativa si existe.
        - Las subcategorías se exportan separadas por comas.

        Args:
            queryset: QuerySet de productos (default: todos activos)
            file_path: ruta de destino o None para generar en memoria

        Returns:
            ruta absoluta del archivo generado o BytesIO si file_path is None
        """
        import io
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        if queryset is None:
            queryset = Product.objects.select_related(
                'category'
            ).prefetch_related('subcategories', 'images').all()

        # Orden exacto de columnas según template web
        WEB_COLUMNS = [
            'code', 'price', 'name', 'diameter', 'length', 'description',
            'images', 'stock', 'category', 'subcategories', 'brand',
            'condition', 'gallery', 'norm', 'grade', 'material', 'colour',
            'type', 'form', 'thread_formats', 'origin', 'faq', 'gtin',
            'mpn', 'meta_title', 'meta_description', 'meta_keywords',
            'google_category',
        ]

        data = []
        for p in queryset:
            # Nombre base sin dimensiones (la web las agrega desde diameter + length)
            base_name = p.get_base_name()

            # Imagen principal
            main_img = ''
            if p.main_image:
                try:
                    main_img = os.path.basename(p.main_image.name)
                except (AttributeError, ValueError):
                    main_img = ''

            # Galería: imágenes adicionales separadas por coma
            gallery_imgs = ', '.join(
                os.path.basename(img.image.name)
                for img in p.images.all()
                if img.image and hasattr(img.image, 'name')
            )

            # FAQs: recopilar de subcategorías del producto
            # Formato web: "¿Pregunta?" Respuesta, "¿Pregunta2?" Respuesta2
            faq_parts = []
            for subcat in p.subcategories.all():
                if subcat.faqs:
                    for faq_item in subcat.faqs:
                        q = faq_item.get('question', '')
                        a = faq_item.get('answer', '')
                        if q and a:
                            faq_parts.append(f'"{q}" {a}')
            faq_str = ', '.join(faq_parts)

            data.append({
                'code': str(p.code),
                'price': float(p.price),
                'name': base_name,
                'diameter': p.diameter or '',
                'length': p.length or '',
                'description': p.description or '',
                'images': main_img,
                'stock': p.stock_quantity,
                'category': getattr(p.category, 'name', '') if p.category else '',
                'subcategories': ', '.join(
                    p.subcategories.values_list('name', flat=True)
                ),
                'brand': p.brand or 'Bulonera Alvear',
                'condition': p.condition or 'new',
                'gallery': gallery_imgs,
                'norm': p.norm or '',
                'grade': p.grade or '',
                'material': p.material or '',
                'colour': p.colour or '',
                'type': p.product_type or '',
                'form': p.form or '',
                'thread_formats': p.thread_format or '',
                'origin': p.origin or '',
                'other_codes': p.other_codes or '',
                'faq': faq_str,
                'gtin': p.gtin or '',
                'mpn': p.mpn or '',
                'meta_title': p.meta_title or '',
                'meta_description': p.meta_description or '',
                'meta_keywords': p.meta_keywords or '',
                'google_category': p.google_category or '',
            })

        # Crear DataFrame con orden exacto de columnas
        df = pd.DataFrame(data, columns=WEB_COLUMNS)

        temp_buffer = io.BytesIO()
        df.to_excel(temp_buffer, index=False, engine='openpyxl')
        temp_buffer.seek(0)

        # ── Aplicar formato de colores al header ────────────────────
        wb = load_workbook(temp_buffer)
        ws = wb.active

        # Colores según política web
        GREEN_FILL = PatternFill(
            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
        )
        RED_FILL = PatternFill(
            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
        )
        BLUE_FILL = PatternFill(
            start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'
        )
        SPEC_FILL = PatternFill(
            start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
        )

        HEADER_FONT = Font(bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        OBLIGATORIAS = {'code', 'price'}
        SEO = {'meta_title', 'meta_description', 'meta_keywords',
               'google_category'}
        SPECS = {'norm', 'grade', 'material', 'colour', 'type', 'form',
                 'thread_formats', 'origin'}

        for col_idx, cell in enumerate(ws[1], 1):
            col_name = cell.value
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

            if col_name in OBLIGATORIAS:
                cell.fill = GREEN_FILL
            elif col_name in SEO:
                cell.fill = BLUE_FILL
            elif col_name in SPECS:
                cell.fill = SPEC_FILL
            else:
                cell.fill = RED_FILL

            # Ajustar ancho de columna
            max_length = len(str(col_name)) + 4
            ws.column_dimensions[cell.column_letter].width = max(
                max_length, 14
            )

        if file_path is None:
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            output_buffer.seek(0)
            logger.info(f"Productos exportados para web en memoria ({len(data)} productos)")
            return output_buffer

        wb.save(file_path)

        logger.info(
            f"Productos exportados para web: {file_path} "
            f"({len(data)} productos)"
        )
        return file_path
