import pytest
import json
from django.urls import reverse
from openpyxl import load_workbook
from io import BytesIO

@pytest.mark.django_db
class TestDashboardPeriodFilter:
    """Pruebas de filtrado por período y gráficos en el dashboard."""

    def test_dashboard_period_today(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=today"
        response = web_client.get(url)
        assert response.status_code == 200
        assert 'chart_labels' in response.context
        assert 'chart_counts' in response.context
        assert 'chart_totals' in response.context

    def test_dashboard_period_week(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=week"
        response = web_client.get(url)
        assert response.status_code == 200

    def test_dashboard_period_month(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=month"
        response = web_client.get(url)
        assert response.status_code == 200

    def test_dashboard_period_quarter(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=quarter"
        response = web_client.get(url)
        assert response.status_code == 200

    def test_dashboard_period_year(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=year"
        response = web_client.get(url)
        assert response.status_code == 200

    def test_dashboard_period_custom(self, web_client):
        url = reverse('sales_web:dashboard') + "?period=custom&from=2026-07-01&to=2026-07-15"
        response = web_client.get(url)
        assert response.status_code == 200
        assert response.context['date_from'] == '2026-07-01'
        assert response.context['date_to'] == '2026-07-15'

    def test_dashboard_period_invalid_custom(self, web_client):
        # Fechas inválidas deben caer en default de mes actual
        url = reverse('sales_web:dashboard') + "?period=custom&from=not-a-date&to=invalid"
        response = web_client.get(url)
        assert response.status_code == 200
        assert "Este Mes" in response.context['period_label']


@pytest.mark.django_db
class TestSalesExcelExport:
    """Pruebas de la exportación a Excel de Ventas."""

    def test_sale_export_requires_login(self, client):
        url = reverse('sales_web:sale_list_export')
        response = client.get(url)
        assert response.status_code == 302

    def test_sale_export_summary_xlsx(self, web_client, sale):
        url = reverse('sales_web:sale_list_export')
        response = web_client.get(url)
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        assert 'Content-Disposition' in response
        assert 'Ventas_' in response['Content-Disposition']

        # Verificar que es un Excel válido leyendo el contenido
        wb = load_workbook(BytesIO(response.content))
        assert "Ventas" in wb.sheetnames
        ws = wb["Ventas"]
        # Fila 1 a 3 tienen metadata, Fila 5 tiene headers, Fila 6 tiene el primer item
        assert ws['A1'].value == "BULONERA ALVEAR — ERP"
        assert ws['A5'].value == "Nro. Venta"
        assert ws['A6'].value == sale.number

    def test_sale_export_detailed_xlsx(self, web_client, sale_with_items):
        url = reverse('sales_web:sale_list_export') + "?detail=1"
        response = web_client.get(url)
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Ventas"]
        assert ws['G5'].value == "Código Item"
        assert ws['H5'].value == "Producto"
        # Primer item detallado de la venta
        assert ws['A6'].value == sale_with_items.number
        assert ws['G6'].value == sale_with_items.items.first().product.code

    def test_sale_export_filters(self, web_client, sale):
        # Filtrar por un número que no existe
        url = reverse('sales_web:sale_list_export') + "?search=NON_EXISTENT_NUMBER"
        response = web_client.get(url)
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Ventas"]
        # No debe haber datos de ventas en la fila 6
        assert ws['A6'].value is None


@pytest.mark.django_db
class TestQuotesExcelExport:
    """Pruebas de la exportación a Excel de Presupuestos."""

    def test_quote_export_requires_login(self, client):
        url = reverse('sales_web:quote_list_export')
        response = client.get(url)
        assert response.status_code == 302

    def test_quote_export_summary_xlsx(self, web_client, quote):
        url = reverse('sales_web:quote_list_export')
        response = web_client.get(url)
        assert response.status_code == 200
        assert response['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        wb = load_workbook(BytesIO(response.content))
        assert "Presupuestos" in wb.sheetnames
        ws = wb["Presupuestos"]
        assert ws['A1'].value == "BULONERA ALVEAR — ERP"
        assert ws['A5'].value == "Nro. Presupuesto"
        assert ws['A6'].value == quote.number

    def test_quote_export_detailed_xlsx(self, web_client, quote, product):
        from sales.models import QuoteItem
        QuoteItem.objects.create(
            quote=quote,
            product=product,
            quantity=5,
            unit_price=product.price,
            tax_percentage=21
        )
        url = reverse('sales_web:quote_list_export') + "?detail=1"
        response = web_client.get(url)
        assert response.status_code == 200
        
        wb = load_workbook(BytesIO(response.content))
        ws = wb["Presupuestos"]
        assert ws['F5'].value == "Código Item"
        assert ws['G5'].value == "Producto"
        assert ws['A6'].value == quote.number
        assert ws['F6'].value == product.code
