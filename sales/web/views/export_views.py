# sales/web/views/export_views.py
"""
Vistas web para la exportación de ventas y presupuestos a Excel.
Maneja los mismos filtros aplicados en el listado y genera un reporte en openpyxl.
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db import models

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sales.models import Sale, Quote
from sales.web.views.web_views import _is_privileged

# Estilos reutilizables para openpyxl
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1B3A5C")
BORDER_SIDE = Side(style='thin', color='D1D5DB')
BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

def _style_worksheet(ws):
    """Auto-ajusta el ancho de las columnas y aplica bordes."""
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format == '$#,##0.00' and isinstance(cell.value, (int, float, Decimal)):
                val_str = f"${float(cell.value):,.2f}"
            max_len = max(max_len, len(val_str))
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)


@login_required
def sale_list_export(request):
    """
    Exporta el listado de ventas filtrado a Excel.
    Soporta ?detail=1 para incluir los items de cada venta.
    """
    sales = Sale.objects.select_related('customer', 'created_by').prefetch_related('items__product').order_by('-date')

    # Restringir según privilegios (operators solo ven lo propio)
    if not _is_privileged(request.user):
        sales = sales.filter(created_by=request.user)

    # Aplicar los mismos filtros que en sale_list
    status_filter  = request.GET.get('status', '').strip()
    payment_filter = request.GET.get('payment_status', '').strip()
    search         = request.GET.get('search', '').strip()

    if status_filter:
        sales = sales.filter(status=status_filter)
    if payment_filter:
        sales = sales.filter(payment_status=payment_filter)
    if search:
        sales = sales.filter(
            models.Q(number__icontains=search)
            | models.Q(customer__business_name__icontains=search)
            | models.Q(items__product__code__icontains=search)
            | models.Q(items__product__sku__icontains=search)
            | models.Q(items__product__other_codes__icontains=search)
        ).distinct()

    # Generar Libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.views.sheetView[0].showGridLines = True

    detail_mode = request.GET.get('detail') == '1'
    
    # Título del Reporte
    ws['A1'] = "BULONERA ALVEAR — ERP"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Reporte de Ventas ({'Detallado' if detail_mode else 'Resumido'})"
    ws['A2'].font = Font(name="Arial", italic=True, size=11)
    ws['A3'] = f"Generado por: {request.user.get_full_name() or request.user.username} | Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(name="Arial", size=9, color="555555")

    start_row = 5

    if not detail_mode:
        # Modo Resumen
        headers = [
            "Nro. Venta", "Fecha", "Cliente", "Estado Venta", "Estado Pago", 
            "Medio Pago", "Creado Por", "Subtotal", "Descuento", "IVA", "Total"
        ]
        ws.row_dimensions[start_row].height = 24
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        curr_row = start_row + 1
        for sale in sales:
            row_values = [
                sale.number,
                sale.date.strftime('%Y-%m-%d %H:%M') if sale.date else '',
                sale.customer.business_name if sale.customer else (sale.customer_name or "Consumidor Final"),
                sale.get_status_display(),
                sale.get_payment_status_display(),
                sale.get_payment_method_display() if sale.payment_method else 'N/A',
                sale.created_by.get_full_name() or sale.created_by.username if sale.created_by else 'Sistema',
                sale._cached_subtotal,
                sale._cached_discount,
                sale._cached_tax,
                sale._cached_total
            ]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = BORDER
                if col_idx in [1, 2, 4, 5, 6, 7]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [8, 9, 10, 11]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '$#,##0.00'
            curr_row += 1

    else:
        # Modo Detallado
        headers = [
            "Nro. Venta", "Fecha", "Cliente", "Estado Venta", "Estado Pago", "Creado Por",
            "Código Item", "Producto", "Cant.", "Precio Unit.", "Tipo Desc.", "Val. Desc.", 
            "IVA %", "Subtotal Línea", "Dcto Línea", "IVA Línea", "Total Línea"
        ]
        ws.row_dimensions[start_row].height = 24
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        curr_row = start_row + 1
        for sale in sales:
            for item in sale.items.all():
                row_values = [
                    sale.number,
                    sale.date.strftime('%Y-%m-%d %H:%M') if sale.date else '',
                    sale.customer.business_name if sale.customer else (sale.customer_name or "Consumidor Final"),
                    sale.get_status_display(),
                    sale.get_payment_status_display(),
                    sale.created_by.username if sale.created_by else 'Sistema',
                    item.product.code,
                    item.product.name,
                    item.quantity,
                    item.unit_price,
                    item.get_discount_type_display(),
                    item.discount_value,
                    item.tax_percentage,
                    item.line_subtotal,
                    item.discount_amount,
                    item.tax_amount,
                    item.total
                ]
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=curr_row, column=col_idx, value=val)
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    if col_idx in [1, 2, 4, 5, 6, 7, 11, 13]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in [9, 10, 12, 14, 15, 16, 17]:
                        cell.alignment = Alignment(horizontal="right")
                        if col_idx in [10, 12, 14, 15, 16, 17]:
                            cell.number_format = '$#,##0.00'
                        elif col_idx == 9:
                            cell.number_format = '#,##0.00'
                curr_row += 1

    _style_worksheet(ws)

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Ventas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    return response


@login_required
def quote_list_export(request):
    """
    Exporta el listado de presupuestos filtrado a Excel.
    Soporta ?detail=1 para incluir los items de cada presupuesto.
    """
    quotes = Quote.objects.select_related('customer', 'created_by').prefetch_related('items__product').order_by('-date')

    # Restringir según privilegios
    if not _is_privileged(request.user):
        quotes = quotes.filter(created_by=request.user)

    # Aplicar los mismos filtros que en quote_list
    status_filter = request.GET.get('status', '').strip()
    search        = request.GET.get('search', '').strip()

    if status_filter:
        quotes = quotes.filter(status=status_filter)
    if search:
        quotes = quotes.filter(
            models.Q(number__icontains=search)
            | models.Q(customer__business_name__icontains=search)
            | models.Q(items__product__code__icontains=search)
            | models.Q(items__product__sku__icontains=search)
            | models.Q(items__product__other_codes__icontains=search)
        ).distinct()

    # Generar Libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Presupuestos"
    ws.views.sheetView[0].showGridLines = True

    detail_mode = request.GET.get('detail') == '1'
    
    # Título del Reporte
    ws['A1'] = "BULONERA ALVEAR — ERP"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Reporte de Presupuestos ({'Detallado' if detail_mode else 'Resumido'})"
    ws['A2'].font = Font(name="Arial", italic=True, size=11)
    ws['A3'] = f"Generado por: {request.user.get_full_name() or request.user.username} | Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'].font = Font(name="Arial", size=9, color="555555")

    start_row = 5

    if not detail_mode:
        # Modo Resumen
        headers = [
            "Nro. Presupuesto", "Fecha", "Válido Hasta", "Cliente", "Estado", 
            "Impreso", "Compartido WA", "Enviado Email", "Creado Por", 
            "Subtotal", "Descuento", "IVA", "Total"
        ]
        ws.row_dimensions[start_row].height = 24
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        curr_row = start_row + 1
        for quote in quotes:
            row_values = [
                quote.number,
                quote.date.strftime('%Y-%m-%d') if quote.date else '',
                quote.valid_until.strftime('%Y-%m-%d') if quote.valid_until else '',
                quote.customer.business_name if quote.customer else (quote.customer_name or "Consumidor Final"),
                quote.get_status_display(),
                "Sí" if quote.is_printed else "No",
                "Sí" if quote.sent_via_wa else "No",
                "Sí" if quote.sent_via_email else "No",
                quote.created_by.get_full_name() or quote.created_by.username if quote.created_by else 'Sistema',
                quote._cached_subtotal,
                quote._cached_discount,
                quote._cached_tax,
                quote._cached_total
            ]
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=curr_row, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.border = BORDER
                if col_idx in [1, 2, 3, 5, 6, 7, 8, 9]:
                    cell.alignment = Alignment(horizontal="center")
                elif col_idx in [10, 11, 12, 13]:
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '$#,##0.00'
            curr_row += 1

    else:
        # Modo Detallado
        headers = [
            "Nro. Presupuesto", "Fecha", "Cliente", "Estado", "Creado Por",
            "Código Item", "Producto", "Cant.", "Precio Unit.", "Tipo Desc.", "Val. Desc.", 
            "IVA %", "Subtotal Línea", "Dcto Línea", "IVA Línea", "Total Línea"
        ]
        ws.row_dimensions[start_row].height = 24
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER

        curr_row = start_row + 1
        for quote in quotes:
            for item in quote.items.all():
                row_values = [
                    quote.number,
                    quote.date.strftime('%Y-%m-%d') if quote.date else '',
                    quote.customer.business_name if quote.customer else (quote.customer_name or "Consumidor Final"),
                    quote.get_status_display(),
                    quote.created_by.username if quote.created_by else 'Sistema',
                    item.product.code,
                    item.product.name,
                    item.quantity,
                    item.unit_price,
                    item.get_discount_type_display(),
                    item.discount_value,
                    item.tax_percentage,
                    item.line_subtotal,
                    item.discount_amount,
                    item.tax_amount,
                    item.total
                ]
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=curr_row, column=col_idx, value=val)
                    cell.font = DATA_FONT
                    cell.border = BORDER
                    if col_idx in [1, 2, 4, 5, 6, 10, 12]:
                        cell.alignment = Alignment(horizontal="center")
                    elif col_idx in [8, 9, 11, 13, 14, 15, 16]:
                        cell.alignment = Alignment(horizontal="right")
                        if col_idx in [9, 11, 13, 14, 15, 16]:
                            cell.number_format = '$#,##0.00'
                        elif col_idx == 8:
                            cell.number_format = '#,##0.00'
                curr_row += 1

    _style_worksheet(ws)

    output = BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Presupuestos_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    return response
